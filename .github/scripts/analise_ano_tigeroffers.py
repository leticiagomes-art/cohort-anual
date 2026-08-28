#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TigerOffers - Analise anual: Pedidos x Reembolsos x Chargebacks
================================================================
Le TODOS os arquivos da pasta de entrada (xlsx/xls/csv/tsv, com N abas por
produto), classifica cada aba em PEDIDOS / REEMBOLSOS / CHARGEBACKS pela
assinatura de colunas, valida a base e gera:

  - Resumo do ano  (geral + por produto)
  - Mensal         (base caixa = data do evento)
  - Mensal coorte  (base coorte = data do PEDIDO que gerou o evento)
  - Semanal
  - Cohort semanal por mes  ($ e %)   -> semana da compra x W+0..W+12
  - Decisor x Mes / Decisor x Produto (System, HelpGrid, Agente BuyGoods,
                                       Agente Tiger)
  - Fases de CS   (HelpGrid -> centralizado -> CS Tiger no helpdesk BuyGoods)
  - Helpdesk 17/08+ (detalhe diario do periodo novo)
  - Afiliados (externos) e Contas Internas (separado)
  - Motivos, Duplicidades, Cobertura & Validacao, LEIA-ME

Uso:
    python analise_ano_tigeroffers.py
    python analise_ano_tigeroffers.py --entrada PASTA --saida PASTA --ano 2026

Nada e inventado: toda coluna que nao for reconhecida aparece no relatorio de
inventario, e toda metrica derivada (%) e suprimida quando o denominador esta
incompleto.
"""

import argparse
import os
import re
import sys
import unicodedata
import warnings
from datetime import date, datetime, timedelta

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

# =====================================================================
# 1. CONFIGURACAO  (unico bloco que precisa ser editado)
# =====================================================================

ENTRADA_PADRAO = "/mnt/user-data/uploads"
SAIDA_PADRAO = "/mnt/user-data/outputs"
ANO_PADRAO = 2026

# Fases da operacao de atendimento. (data_inicio, rotulo)
# 17/08/2026 = inicio do atendimento da Tiger no helpdesk da BuyGoods.
FASES = [
    (date(2000, 1, 1), "1 - HelpGrid ativa (ligacao + tickets)"),
    (date(2026, 5, 1), "2 - Centralizado interno por e-mail / automatico BuyGoods"),
    (date(2026, 8, 17), "3 - CS Tiger no helpdesk da BuyGoods"),
]
DATA_HELPDESK_INTERNO = date(2026, 8, 17)

# Somente estes sao agentes da Tiger. Qualquer outra sigla e agente da BuyGoods.
AGENTES_TIGER = {"mari alves", "marialves", "leticia gomes", "leticiagomes",
                 "mari", "leticia"}

# Contas internas / plataforma: ficam FORA do ranking de afiliados,
# mas continuam dentro dos totais gerais.
CONTAS_INTERNAS = [
    "helpgrid", "help grid", "maxweb", "maxweb 2", "mwe", "mwe ltda",
    "tiger offers ltda", "tigeroffers ltda", "tiger offers", "tigeroffers",
    "jose moraes", "josemoraes", "gestor1.buygoods", "gestor1buygoods",
    "gestor3buygoods", "gestor3.buygoods", "gestor five", "gestorfive",
    "gestor5", "gestor 5",
]

# Flags da base de Order Items que EXCLUEM a linha do Gross.
# Padrao: so 'Was Declined' (pagamento recusado, dinheiro nunca capturado).
# Cancelados e voids ficam no Gross e sao reportados na aba de validacao.
EXCLUIR_DE_GROSS = ["was_declined"]

# Virada de funil Black -> White (fonte: analise_black_white_produtos).
# Antes da data o produto rodava MIX 70% Black / 30% White; a partir dela,
# 100% White - excecao BreathEaseX, que passou a 50/50.
VIRADA_BLACK_WHITE = {
    "ReduTide": "2026-07-07", "NervoLyn": "2026-07-24", "Prostafense": "2026-07-27",
    "AudiLeaf": "2026-07-28", "VisiumPro": "2026-07-28", "MaroBrain": "2026-07-28",
    "GlucoRecover": "2026-07-31", "FloraNew": "2026-07-31", "AlphaErec": "2026-08-04",
    "BoosterXT": "2026-08-07", "NailsCleanPro": "2026-08-12",
    "BreathEaseX": "2026-08-12", "LipoBliss": "2026-08-13", "LipoPeak": "2026-08-14",
    "MounjaMelt": "2026-08-14",
}
OBS_VIRADA = {"BreathEaseX": "passou a 50/50, nao a 100% White"}

# Horizonte do cohort semanal (semanas apos a compra)
COHORT_MAX_SEMANAS = 12

# Cobertura minima de dias com pedido para liberar calculo de % no mes
COBERTURA_MINIMA = 0.90

# Escala de alerta usada no morning brief (refund + CB sobre gross)
ESCALA_ALERTA = [(0.30, "VERMELHO >=30%"), (0.20, "LARANJA 20-29%"),
                 (0.10, "AMARELO 10-19%"), (0.00, "OK <10%")]

# Paleta executiva
NAVY = "1F3864"
BLUE = "2F5597"
CINZA = "F2F2F2"
BRANCO = "FFFFFF"
VERDE = "C6EFCE"
VERMELHO = "FFC7CE"
AMARELO = "FFEB9C"

# =====================================================================
# 2. DICIONARIO DE COLUNAS
# =====================================================================

def norm(s):
    """minusculo, sem acento, so alfanumerico.

    A remocao de acento e obrigatoria: sem ela 'Leticia Gomes' e
    'Leticia Gomes' com acento viram chaves diferentes e o agente do time
    e classificado como agente da BuyGoods.
    """
    t = unicodedata.normalize("NFKD", str(s).strip().lower())
    t = t.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", t)


# canonico -> lista de aliases (ja normalizados). Casamento e EXATO de
# proposito: 'External Order ID' nunca pode ser lido como 'Order ID'.
ALIASES = {
    "order_id":        ["orderid", "order", "idpedido", "pedidoid"],
    "order_item_id":   ["orderitemid", "itemid", "orderitem"],
    "account_id":      ["accountid", "contaid"],
    "user_id":         ["userid"],
    "affiliate_id":    ["affiliateid", "affid", "idafiliado", "affiliateidno"],
    "affiliate_name":  ["affiliatename", "affiliate", "affiliateusername",
                        "username", "affiliatelogin", "nomeafiliado"],
    "product":         ["productname", "product", "produto", "productcodename",
                        "productcode", "offername"],
    "product_codename": ["productcodename", "productcode", "codename"],
    "tipo":            ["type", "refundtype", "transactiontype", "tipo"],
    "is_recurring":    ["isrecurring", "recurring", "rebill", "isrebill"],
    "no_charges":      ["noofcharges", "numberofcharges", "nofcharges",
                        "chargesno", "charges"],
    "commission_amount": ["commissionamount", "affiliatecommission", "commission"],
    "amount":          ["amount", "totalamount", "ordertotal", "total",
                        "refundamount", "amountrefunded", "chargebackamount",
                        "saleamount", "grossamount", "value", "valor",
                        "price", "orderamount", "transactionamount"],
    "quantity":        ["quantity", "qty", "quantidade", "units"],
    "order_date":      ["orderdate", "purchasedate", "saledate", "datapedido"],
    "date_created":    ["datecreated", "createdat", "createddate", "datacriacao"],
    "refund_date":     ["refunddate", "datereembolso", "datarefund",
                        "refundedat", "refundedon"],
    "void_date":       ["voiddate"],
    "cb_date":         ["chargebackdate", "cbdate", "disputedate",
                        "chargebackcreated", "datechargeback"],
    "fulfillment_date": ["fulfillmentdate", "shippeddate", "shipdate"],
    "was_fulfilled":   ["wasfulfilled", "fulfilled", "isfulfilled"],
    "agent":           ["agentname", "agent", "refundedby", "processedby",
                        "agente", "createdby", "handledby", "operator"],
    # 'Comments' e 'Notes' ficam DE FORA de proposito: o export de Order Items
    # tem uma coluna 'Comments' que nao e motivo de reembolso.
    "reason":          ["reason", "refundreason", "reasoncode", "motivo",
                        "chargebackreason", "reasondescription"],
    "customer_email":  ["customeremail", "customeremailaddress", "email",
                        "buyeremail", "emailaddress"],
    "customer_name":   ["customername", "buyername", "name"],
    "customer_country": ["customercountry", "country", "pais"],
    "customer_first":  ["firstname", "first"],
    "customer_last":   ["lastname", "last"],
    "phone":           ["phone", "phonenumber", "telefone", "customerphone"],
    "state":           ["state", "estado", "customerstate"],
    "city":            ["city", "cidade", "customercity"],
    "zip":             ["zip", "zipcode", "customerzipcode", "postalcode"],
    "referrer":        ["referrerurl", "referrer", "referrerlink"],
    "categoria":       ["categoryname", "category", "nicho"],
    "was_canceled":    ["wascanceled", "wascancelled"],
    "was_refunded":    ["wasrefunded"],
    "was_chargeback":  ["waschargeback", "waschargebacked"],
    "was_voided":      ["wasvoided"],
    "was_declined":    ["wasdeclined"],
    "cancel_date":     ["canceldate"],
    "cancel_reason":   ["cancelreason"],
    "declined_date":   ["declineddate"],
    "shipping_cost":   ["shippingcost"],
    "insurance_cost":  ["insurancecost"],
    "taxes":           ["taxes", "tax"],
    "external_fees":   ["externalfees"],
    "provider_cost":   ["providercost"],
    "store":           ["storename", "store"],
    "next_charge_date": ["nextchargedate"],
    "last_access":     ["lastaccess"],
    "status":          ["status", "orderstatus", "refundstatus"],
}

# prioridade quando ha varios candidatos para o mesmo canonico
PRIORIDADE_AMOUNT = ["refundamount", "amountrefunded", "chargebackamount",
                     "amount", "totalamount", "ordertotal", "orderamount",
                     "transactionamount", "saleamount", "grossamount",
                     "total", "value", "valor", "price"]
# "Commission Amount" NAO entra em amount: e a comissao devolvida ao afiliado,
# nao o valor reembolsado ao cliente.

PRODUTOS_CANON = [
    "BreathEaseX", "CardioEaseX", "NervoLyn", "AudiLeaf", "VisiumPro",
    "MaroBrain", "Prostafense", "VigorLong", "GlucoRecover", "FloraNew",
    "BoosterXT", "LipoBliss", "LipoPeak", "LipoVive", "ReduTide", "ReduBurn",
    "Lipotrine", "MounjaMelt", "NailsCleanPro", "AlphaErec", "Nuraliss",
    "Erectrozil", "IronPulseX",
]
MAPA_PRODUTO = {norm(p): p for p in PRODUTOS_CANON}

# Alguns exports trazem o NOME DA OFERTA em Product Name ("3 Bottles of X + ...")
# em vez do produto. Nesses casos o Product Codename tem um prefixo estavel.
# Grafias divergentes entre a aba do workbook e a conta no Master oficial
ALIAS_MASTER = {"redurburn": "reduburn", "oaztem": "oatzem",
                "reduburn": "redurburn", "oatzem": "oaztem"}

PREFIXO_CODENAME = {
    "breex": "BreathEaseX", "nervo": "NervoLyn", "audi": "AudiLeaf",
    "prosta": "Prostafense", "visium": "VisiumPro", "flora": "FloraNew",
    "nails": "NailsCleanPro", "maro": "MaroBrain", "gluco": "GlucoRecover",
    "booster": "BoosterXT", "vigor": "VigorLong", "lipo": "LipoBliss",
    "redu": "ReduBurn", "iron": "IronPulseX", "erect": "Erectrozil",
}


def produto_por_texto(texto):
    """Acha um produto conhecido dentro de um nome de oferta ou codename."""
    n = norm(texto)
    if not n:
        return None
    for k, v in MAPA_PRODUTO.items():
        if k and k in n:
            return v
    for pref, v in PREFIXO_CODENAME.items():
        if n.startswith(pref):
            return v
    return None

LIXO_DATA = {"", "0000-00-00", "0000-00-00 00:00:00", "0000-00-00 00:00",
             "nan", "nat", "none", "null", "-", "n/a", "na"}


# =====================================================================
# 3. LEITURA E PARSING
# =====================================================================

def canonizar_produto(nome):
    n = norm(nome)
    if n in MAPA_PRODUTO:
        return MAPA_PRODUTO[n]
    for k, v in MAPA_PRODUTO.items():
        if k and (k in n or n in k):
            return v
    return str(nome).strip()


def parse_data(serie):
    """Parser de data tolerante: ISO, 'August 26, 2026', 0000-00-00 -> NaT."""
    if serie is None:
        return pd.Series(pd.NaT, index=[])
    if pd.api.types.is_datetime64_any_dtype(serie):
        return pd.to_datetime(serie, errors="coerce")
    s = serie.astype(str).str.strip()
    s = s.mask(s.str.lower().isin(LIXO_DATA), None)
    s = s.mask(s.str.startswith("0000-00-00", na=False), None)
    try:
        out = pd.to_datetime(s, errors="coerce", format="mixed", dayfirst=False)
    except Exception:
        out = pd.to_datetime(s, errors="coerce", dayfirst=False)
    faltando = out.isna() & s.notna()
    if faltando.any():
        alt = pd.to_datetime(s[faltando], errors="coerce", dayfirst=False)
        out.loc[faltando] = alt
    return out


def _num_escalar(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return np.nan
    t = str(x).strip()
    if t == "" or t.lower() in LIXO_DATA:
        return np.nan
    neg = t.startswith("(") and t.endswith(")")
    t = re.sub(r"[^0-9,.\-]", "", t)
    if t in ("", "-", ".", ","):
        return np.nan
    if "," in t and "." in t:
        if t.rfind(",") > t.rfind("."):
            t = t.replace(".", "").replace(",", ".")
        else:
            t = t.replace(",", "")
    elif "," in t:
        # virgula unica com 1-2 casas = decimal; senao e milhar
        if re.search(r",\d{1,2}$", t):
            t = t.replace(",", ".")
        else:
            t = t.replace(",", "")
    try:
        v = float(t)
    except ValueError:
        return np.nan
    return -v if neg else v


def parse_valor(serie):
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce")
    return serie.map(_num_escalar)


def detectar_cabecalho(df_raw, limite=12):
    """Acha a linha real do cabecalho (exports BuyGoods tem 3-4 linhas de meta)."""
    alvos = {"orderid", "orderitemid", "affiliateid", "refunddate", "orderdate"}
    for i in range(min(limite, len(df_raw))):
        celulas = {norm(c) for c in df_raw.iloc[i].tolist() if str(c) != "nan"}
        if len(celulas & alvos) >= 2:
            return i
    return 0


def mapear_colunas(cols):
    """Retorna (mapa canonico->coluna original, lista de colunas nao mapeadas)."""
    normalizadas = {}
    for c in cols:
        normalizadas.setdefault(norm(c), []).append(c)
    mapa, usadas = {}, set()
    # amount tem regra de prioridade propria
    for alias in PRIORIDADE_AMOUNT:
        if alias in normalizadas and "amount" not in mapa:
            mapa["amount"] = normalizadas[alias][0]
            usadas.add(normalizadas[alias][0])
    for canon, aliases in ALIASES.items():
        if canon in mapa:
            continue
        for alias in aliases:
            if alias in normalizadas:
                escolha = [c for c in normalizadas[alias] if c not in usadas]
                if escolha:
                    mapa[canon] = escolha[0]
                    usadas.add(escolha[0])
                    break
    nao_mapeadas = [c for c in cols if c not in usadas and str(c).strip() != ""
                    and not str(c).lower().startswith("unnamed")]
    return mapa, nao_mapeadas


def classificar_tipo(mapa, nome_arquivo, nome_aba):
    """PEDIDOS / REEMBOLSOS / CHARGEBACKS / DESCONHECIDO.

    A ordem importa: o export de Order Items TAMBEM tem 'Void Date', entao a
    checagem de PEDIDOS vem antes e void_date sozinho nunca decide nada.
    O export de Chargeback nao tem coluna de data do evento nem Agent Name -
    o unico discriminador contra o de Refund e a ausencia de Refund Date.
    """
    txt = norm(nome_arquivo) + norm(nome_aba)
    if "order_item_id" in mapa or "was_fulfilled" in mapa or "was_refunded" in mapa:
        return "PEDIDOS"
    if "refund_date" in mapa:
        return "REEMBOLSOS"
    if "cb_date" in mapa:
        return "CHARGEBACKS"
    if any(k in txt for k in ("chargeback", "estorno", "cb2026", "cb20")):
        return "CHARGEBACKS"
    if any(k in txt for k in ("reembolso", "refund")):
        return "REEMBOLSOS"
    if any(k in txt for k in ("orderitem", "order", "pedido", "vendas", "sales")):
        return "PEDIDOS"
    return "DESCONHECIDO"


def flag_true(serie):
    """Was Refunded / Was Canceled etc. podem vir como 1/0, Yes/No ou TRUE/FALSE."""
    return serie.astype(str).str.strip().str.lower().isin(
        {"1", "1.0", "true", "yes", "y", "sim", "t", "x"})


def eh_master(caminho):
    """Master Overview / Master Accounts: relatorio agregado oficial da BuyGoods."""
    n = norm(os.path.basename(caminho))
    return n.startswith("masteroverview") or n.startswith("masteraccounts")


def ler_master(caminho):
    """Acha a linha de cabecalho (a que tem 'Gross Sales') e devolve limpo."""
    raw = pd.read_excel(caminho, header=None)
    h = None
    for i in range(min(15, len(raw))):
        if raw.iloc[i].astype(str).str.contains("Gross Sales", na=False).any():
            h = i
            break
    if h is None:
        return pd.DataFrame()
    df = pd.read_excel(caminho, skiprows=h)
    df = df.dropna(how="all").dropna(axis=1, how="all")
    df.columns = [str(c).strip() for c in df.columns]
    chave = df.columns[0]
    df = df[df[chave].notna()]
    df["__total"] = df[chave].astype(str).str.strip().str.lower() == "total"
    return df


def ler_arquivo(caminho):
    """Gera (nome_aba, DataFrame bruto) para cada aba/arquivo."""
    ext = os.path.splitext(caminho)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        sep = "\t" if ext == ".tsv" else None
        bruto = pd.read_csv(caminho, header=None, sep=sep, engine="python",
                            encoding="utf-8-sig", on_bad_lines="skip",
                            dtype=str, nrows=15)
        h = detectar_cabecalho(bruto)
        df = pd.read_csv(caminho, sep=sep, engine="python", skiprows=h,
                         encoding="utf-8-sig", on_bad_lines="skip", dtype=str)
        yield os.path.splitext(os.path.basename(caminho))[0], df
        return
    if ext in (".xlsx", ".xlsm", ".xls"):
        xl = pd.ExcelFile(caminho)
        for aba in xl.sheet_names:
            bruto = pd.read_excel(xl, sheet_name=aba, header=None, nrows=15,
                                  dtype=object)
            if bruto.empty:
                yield aba, pd.DataFrame()
                continue
            h = detectar_cabecalho(bruto)
            df = pd.read_excel(xl, sheet_name=aba, skiprows=h, dtype=object)
            yield aba, df
        return
    raise ValueError("extensao nao suportada: " + ext)


# =====================================================================
# 4. CARGA + INVENTARIO
# =====================================================================

def carregar(entrada):
    inventario, blocos = [], {"PEDIDOS": [], "REEMBOLSOS": [], "CHARGEBACKS": []}
    arquivos = sorted(
        os.path.join(entrada, f) for f in os.listdir(entrada)
        if os.path.splitext(f)[1].lower() in (".xlsx", ".xlsm", ".xls", ".csv",
                                              ".tsv", ".txt")
    )
    if not arquivos:
        print("!! Nenhuma planilha encontrada em", entrada)
        return inventario, blocos

    masters = {}
    for caminho in arquivos:
        nome = os.path.basename(caminho)
        if eh_master(caminho):
            m = ler_master(caminho)
            rot = "OVERVIEW" if norm(nome).startswith("masteroverview") else "ACCOUNTS"
            masters[rot] = m
            inventario.append({"Arquivo": nome, "Aba": "-",
                               "Tipo": "MASTER " + rot, "Linhas": len(m),
                               "Periodo": "relatorio oficial agregado da BuyGoods",
                               "Observacao": "usado so para reconciliacao, nao entra "
                                             "nas bases de detalhe"})
            print("\n== lendo:", nome, "-> MASTER", rot, "(%d linhas)" % len(m))
            continue
        print("\n== lendo:", nome)
        try:
            abas = list(ler_arquivo(caminho))
        except Exception as e:
            inventario.append({"Arquivo": nome, "Aba": "-", "Tipo": "ERRO",
                               "Linhas": 0, "Observacao": "falha ao abrir: %s" % e})
            print("   ERRO ao abrir:", e)
            continue

        for aba, df in abas:
            if df is None or df.empty or len(df.columns) == 0:
                inventario.append({"Arquivo": nome, "Aba": aba, "Tipo": "VAZIA",
                                   "Linhas": 0, "Observacao": "aba sem dados"})
                continue
            df = df.dropna(how="all")
            df = df.loc[:, [c for c in df.columns
                            if not str(c).lower().startswith("unnamed")
                            or df[c].notna().any()]]
            mapa, nao_map = mapear_colunas(df.columns)
            tipo = classificar_tipo(mapa, nome, aba)

            obs = []
            if tipo == "DESCONHECIDO":
                obs.append("NAO CLASSIFICADA - conferir manualmente")
            if "amount" not in mapa:
                obs.append("sem coluna de valor")
            if nao_map:
                obs.append("colunas nao mapeadas: " + ", ".join(map(str, nao_map[:12])))

            std = pd.DataFrame(index=df.index)
            for canon, origem in mapa.items():
                std[canon] = df[origem]
            std["produto"] = canonizar_produto(aba)
            std["arquivo"] = nome
            std["aba"] = aba

            # datas
            for col in ("order_date", "date_created", "refund_date",
                        "void_date", "cb_date", "fulfillment_date",
                        "cancel_date", "declined_date", "next_charge_date"):
                std[col] = parse_data(std[col]) if col in std.columns else pd.NaT
            if "amount" in std.columns:
                std["amount"] = parse_valor(std["amount"])
            else:
                std["amount"] = np.nan
            for col in ("quantity", "commission_amount", "no_charges",
                        "shipping_cost", "insurance_cost", "taxes",
                        "external_fees", "provider_cost"):
                if col in std.columns:
                    std[col] = parse_valor(std[col])

            for col in ("order_id", "affiliate_id", "account_id",
                        "affiliate_name", "agent", "reason", "product",
                        "customer_email", "order_item_id", "status",
                        "tipo", "is_recurring", "product_codename", "state",
                        "customer_first", "customer_last", "referrer",
                        "categoria", "was_canceled", "was_refunded",
                        "was_chargeback", "was_voided", "was_declined",
                        "cancel_reason", "store"):
                if col not in std.columns:
                    std[col] = np.nan
                else:
                    std[col] = std[col].astype(str).str.strip()
                    std[col] = std[col].replace({"nan": np.nan, "None": np.nan,
                                                 "": np.nan})

            # produto: prioriza Product Name da propria linha se existir
            aba_eh_produto = norm(aba) in MAPA_PRODUTO or len(abas) > 1
            if std["product"].notna().any() and not aba_eh_produto:
                brutos = std.loc[std["product"].notna(), "product"]
                distintos = brutos.nunique()
                if distintos == 1:
                    std["produto"] = canonizar_produto(brutos.iloc[0])
                elif distintos > 3:
                    # Product Name virou nome de OFERTA: resolver por token do
                    # produto e, se falhar, pelo codename; o resto herda o
                    # produto dominante do arquivo (e isso vai reportado).
                    res = std["product"].map(produto_por_texto)
                    if "product_codename" in std.columns:
                        res = res.fillna(std["product_codename"].map(produto_por_texto))
                    dominante = (res.mode().iloc[0] if res.notna().any()
                                 else canonizar_produto(aba))
                    n_fall = int(res.isna().sum())
                    std["produto"] = res.fillna(dominante)
                    obs.append("Product Name traz nome de OFERTA (%d valores distintos): "
                               "produto resolvido por token/codename; %d linhas sem match "
                               "herdaram '%s'" % (distintos, n_fall, dominante))
                else:
                    std["produto"] = std["product"].map(canonizar_produto)

            if tipo in blocos:
                blocos[tipo].append(std)

            # periodo detectado
            col_data = {"PEDIDOS": ["date_created", "order_date"],
                        "REEMBOLSOS": ["refund_date"],
                        "CHARGEBACKS": ["cb_date", "order_date"]}.get(tipo, ["order_date"])
            per = ""
            for c in col_data:
                if c in std.columns and std[c].notna().any():
                    per = "%s -> %s (%s)" % (std[c].min().date(),
                                             std[c].max().date(), c)
                    break

            inventario.append({
                "Arquivo": nome, "Aba": aba, "Tipo": tipo, "Linhas": len(std),
                "Produto": std["produto"].iloc[0] if len(std) else "",
                "Periodo": per,
                "Coluna de valor": mapa.get("amount", "AUSENTE"),
                "Colunas reconhecidas": ", ".join(sorted(mapa.keys())),
                "Observacao": " | ".join(obs) if obs else "ok",
            })
            print("   aba %-22s tipo=%-12s linhas=%-7d %s" %
                  (str(aba)[:22], tipo, len(std), per))

    saida = {}
    for tipo, lista in blocos.items():
        saida[tipo] = (pd.concat(lista, ignore_index=True) if lista
                       else pd.DataFrame())
    saida["_MASTERS"] = masters
    return inventario, saida


# =====================================================================
# 5. VALIDACAO
# =====================================================================

def validar(ped, ref, cb, ano, diag=None):
    checks = []

    def add(item, resultado, detalhe=""):
        checks.append({"Verificacao": item, "Resultado": resultado,
                       "Detalhe": detalhe if detalhe else "-"})

    add("Linhas carregadas", "%d pedidos (itens) | %d reembolsos | %d chargebacks"
        % (len(ped), len(ref), len(cb)), "antes de qualquer filtro")

    # 1. Order Date do refund e confiavel?
    if len(ref) and ref["order_date"].notna().any() and ref["refund_date"].notna().any():
        m = ref["order_date"].notna() & ref["refund_date"].notna()
        iguais = (ref.loc[m, "order_date"].dt.date ==
                  ref.loc[m, "refund_date"].dt.date).mean()
        add("Order Date do export de reembolso",
            "%.1f%% das linhas tem Order Date = Refund Date" % (iguais * 100),
            "acima de 50% a coluna e inutilizavel para coorte -> "
            "usar a data do pedido vinda de Orders (join por Order ID)")
    else:
        add("Order Date do export de reembolso", "nao avaliado",
            "coluna ausente ou vazia")

    # 2. taxa de casamento com a base de pedidos
    if len(ped) and ped["order_id"].notna().any():
        ids = set(ped["order_id"].dropna().astype(str))
        for nome, base in (("reembolsos", ref), ("chargebacks", cb)):
            if len(base) and base["order_id"].notna().any():
                taxa = base["order_id"].astype(str).isin(ids).mean()
                add("Casamento %s x base de pedidos" % nome,
                    "%.1f%% dos Order IDs encontrados" % (taxa * 100),
                    "o que nao casa fica sem data real de pedido "
                    "(coorte usa fallback ou vira SEM COORTE)")
    else:
        add("Casamento com base de pedidos", "IMPOSSIVEL",
            "base de pedidos ausente ou sem Order ID")

    # 3. datas nao parseadas
    for nome, base, col in (("pedidos", ped, "date_created"),
                            ("reembolsos", ref, "refund_date"),
                            ("chargebacks", cb, "cb_date")):
        if len(base) and col in base.columns and base[col].notna().any():
            nulos = int(base[col].isna().sum())
            if nulos:
                add("Datas invalidas em %s (%s)" % (nome, col), "%d linhas" % nulos,
                    "linhas sem data ficam fora dos cortes por periodo")

    # 4. valores ausentes
    for nome, base in (("pedidos", ped), ("reembolsos", ref), ("chargebacks", cb)):
        if len(base):
            sem = int(base["amount"].isna().sum())
            if sem:
                add("Valor ausente em %s" % nome, "%d de %d linhas" % (sem, len(base)),
                    "contam no volume, nao no valor")

    # 5. colisao de affiliate_id
    if len(ped) and ped["affiliate_name"].notna().any():
        g = (ped.dropna(subset=["affiliate_id"])
                .groupby(["produto", "affiliate_id"])["affiliate_name"]
                .nunique())
        col = int((g > 1).sum())
        glob = ped.dropna(subset=["affiliate_id"]).groupby(
            "affiliate_id")["affiliate_name"].nunique()
        add("Colisao de Affiliate ID",
            "%d combinacoes produto+id com mais de um nome | %d IDs com mais de um nome "
            "quando se ignora o produto" % (col, int((glob > 1).sum())),
            "aff_id e escopado por conta BuyGoods, nao e global - a identidade usada "
            "nas abas e (produto, ID), e o ranking agrega por nome")

    # 6. duplicidade de reembolso
    dups = pd.DataFrame()
    if len(ref):
        base = ref.dropna(subset=["order_id", "amount"]).copy()
        base["chave"] = base["order_id"].astype(str) + "|" + base["amount"].round(2).astype(str)
        cont = base.groupby("chave").size()
        chaves = cont[cont > 1].index
        if len(chaves):
            dups = base[base["chave"].isin(chaves)].sort_values(["chave", "refund_date"])
            add("Reembolsos duplicados (mesmo Order ID + mesmo valor)",
                "%d linhas em %d pedidos | US$ %s" %
                (len(dups), len(chaves), f"{dups['amount'].sum():,.2f}"),
                "detalhe na aba Duplicidades")
        else:
            add("Reembolsos duplicados", "nenhum", "")

    # 7. legado
    if len(ref) and "data_pedido" in ref.columns:
        leg = int((ref["data_pedido"].dt.year < ano).sum())
        add("Reembolsos de pedidos anteriores a %d (LEGADO)" % ano, "%d linhas" % leg,
            "isolados para nao distorcer a taxa do ano")

    # 7a-1. o que e a coluna Order Date do export de chargeback
    if diag and "cb" in diag:
        mesmo, gap, rot = diag["cb"]
        add("Data usada no chargeback", rot,
            "o export de chargeback nao tem coluna de data do estorno nem Agent Name. "
            "Comparacao com a data real do pedido: %.1f%% no mesmo dia, mediana de %.0f "
            "dias de diferenca." % ((mesmo or 0) * 100, gap if gap == gap else 0))

    # 7a-2. SEGUNDA CHAVE: flags Was Refunded / Was Chargeback da base de pedidos
    for flag, base_ev, rotulo in (("was_refunded", ref, "reembolso"),
                                  ("was_chargeback", cb, "chargeback")):
        if len(ped) and flag in ped.columns and ped[flag].notna().any():
            ids_flag = set(ped.loc[flag_true(ped[flag]), "order_id"].dropna().astype(str))
            ids_base = set(base_ev["order_id"].dropna().astype(str)) if len(base_ev) else set()
            inter = ids_flag & ids_base
            add("Universo de %s por 2 chaves independentes" % rotulo,
                "flag na base de pedidos: %d | arquivo de %s: %d | em ambos: %d"
                % (len(ids_flag), rotulo, len(ids_base), len(inter)),
                "so no flag: %d (evento existe mas nao esta no arquivo) | so no arquivo: "
                "%d (pedido fora da base de Order Items). Gap grande = recorte incompleto, "
                "nao fechar numero por uma chave so."
                % (len(ids_flag - ids_base), len(ids_base - ids_flag)))

    # 7a-3. pedidos cancelados / recusados / void na base de Order Items
    for flag, rotulo in (("was_declined", "recusados (Was Declined)"),
                         ("was_canceled", "cancelados (Was Canceled)"),
                         ("was_voided", "void (Was Voided)")):
        if len(ped) and flag in ped.columns and ped[flag].notna().any():
            m = flag_true(ped[flag])
            if m.any():
                add("Pedidos %s" % rotulo, "%d itens | US$ %s" %
                    (int(m.sum()), f"{ped.loc[m, 'amount'].sum():,.2f}"),
                    "recusado = pagamento nunca capturado, EXCLUIDO do Gross. "
                    "Cancelado e void ficam no Gross e sao reportados aqui para decisao."
                    if flag == "was_declined" else
                    "mantido no Gross; virar exclusao editando EXCLUIR_DE_GROSS no topo do script")

    # 7b. mesmo Affiliate ID com nomes diferentes entre os arquivos
    if len(ref) and "variantes_nome" in ref.columns:
        v = pd.concat([b[["afiliado_id", "variantes_nome"]] for b in (ped, ref, cb)
                       if len(b) and "variantes_nome" in b.columns],
                      ignore_index=True).drop_duplicates("afiliado_id")
        multi = v[v["variantes_nome"].str.contains(r"\|", na=False)]
        add("Affiliate ID com mais de um nome entre os arquivos",
            "%d IDs" % len(multi),
            "; ".join("%s = %s" % (r["afiliado_id"], r["variantes_nome"])
                      for _, r in multi.head(15).iterrows())
            or "nenhum - Refund/CB e Orders usam o mesmo nome")

    # 8. universo real dos campos categoricos (nao presumir valores)
    for campo, rotulo in (("agent", "Agent Name"), ("tipo", "Type"),
                          ("is_recurring", "Is recurring")):
        if len(ref) and campo in ref.columns and ref[campo].notna().any():
            vc = ref[campo].fillna("(vazio)").value_counts()
            if campo == "agent":
                det = "; ".join("%s -> %s (%d)" % (k, classificar_agente(k), v)
                                for k, v in vc.head(20).items())
            else:
                det = "; ".join("%s (%d)" % (k, v) for k, v in vc.head(15).items())
            add("Valores em %s (reembolso)" % rotulo, "%d distintos" % len(vc), det)

    # 9. voids: cancelamento antes da captura nao e reembolso de venda liquidada
    if len(ref) and "void_date" in ref.columns:
        nv = int(ref["void_date"].notna().sum())
        if nv:
            add("Linhas com Void Date preenchido", "%d linhas | US$ %s" %
                (nv, f"{ref.loc[ref['void_date'].notna(), 'amount'].sum():,.2f}"),
                "void = cancelado antes da captura. Mantidos nos totais; separar se a "
                "leitura for de venda liquidada")

    # 10. comissao devolvida
    if len(ref) and "commission_amount" in ref.columns and ref["commission_amount"].notna().any():
        add("Comissao devolvida pelos afiliados",
            "US$ %s" % f"{ref['commission_amount'].sum():,.2f}",
            "coluna Commission Amount do export de reembolso; nao entra no valor reembolsado")

    return pd.DataFrame(checks), dups


def conferencia_fontes(bases, ano):
    """Linha a linha por arquivo/aba: o que foi lido x o que entrou no recorte."""
    linhas = []
    for tipo, base in bases:
        if not len(base):
            continue
        for (arq, aba), g in base.groupby(["arquivo", "aba"]):
            no_ano = g[g["ano"] == ano]
            linhas.append({
                "Tipo": tipo, "Arquivo": arq, "Aba": aba,
                "Linhas lidas": len(g),
                "Valor lido": float(g["amount"].sum(skipna=True)),
                "Linhas no ano %d" % ano: len(no_ano),
                "Valor no ano %d" % ano: float(no_ano["amount"].sum(skipna=True)),
                "Linhas descartadas": len(g) - len(no_ano),
                "Motivo do descarte": ("fora do ano ou sem data"
                                       if len(g) != len(no_ano) else "-"),
            })
    return pd.DataFrame(linhas)


# =====================================================================
# 6. ENRIQUECIMENTO
# =====================================================================

def classificar_agente(nome):
    n = norm(nome)
    if n in ("", "nan", "none", "null"):
        return "System (automatico BuyGoods)"
    if n in ("system", "systemautomatic", "auto", "automatic", "automated"):
        return "System (automatico BuyGoods)"
    if "helpgrid" in n:
        return "HelpGrid"
    if any(norm(a) == n or norm(a) in n for a in AGENTES_TIGER):
        return "Agente Tiger (CS interno)"
    return "Agente BuyGoods"


def fase_de(d):
    if pd.isna(d):
        return "sem data"
    dd = d.date() if hasattr(d, "date") else d
    rot = FASES[0][1]
    for inicio, rotulo in FASES:
        if dd >= inicio:
            rot = rotulo
    return rot


def semana_inicio(s):
    """Segunda-feira da semana, zerada na meia-noite (senao cada linha vira
    uma 'semana' propria por causa da hora)."""
    s = pd.to_datetime(s, errors="coerce")
    return (s - pd.to_timedelta(s.dt.weekday, unit="D")).dt.normalize()


def rotulo_semana(ts):
    if pd.isna(ts):
        return "sem data"
    fim = ts + timedelta(days=6)
    return "%d-W%02d (%s a %s)" % (ts.isocalendar()[0], ts.isocalendar()[1],
                                   ts.strftime("%d/%m"), fim.strftime("%d/%m"))


def eh_interna(nome, aff_id):
    n = norm(nome)
    if n in ("", "nan", "none"):
        return False
    return any(norm(c) in n or n in norm(c) for c in CONTAS_INTERNAS)


def nome_canonico(bases):
    """Um nome por (PRODUTO, Affiliate ID).

    O aff_id da BuyGoods e escopado por CONTA, nao e global: o id 15 pode ser
    pessoas diferentes em produtos diferentes. Resolver o nome dentro do
    produto e so depois agregar por nome evita fundir gente que nao tem nada
    a ver. Refund/CB trazem o USERNAME, Orders traz o NOME CADASTRADO - por
    isso Orders tem peso maior.
    """
    frames = []
    for base, peso in bases:
        if len(base):
            t = base[["produto", "afiliado_id", "afiliado_nome"]].copy()
            t["peso"] = peso
            frames.append(t)
    if not frames:
        return {}, {}
    t = pd.concat(frames, ignore_index=True)
    t = t[t["afiliado_nome"].notna() & (t["afiliado_nome"] != "(sem nome)")]
    if not len(t):
        return {}, {}
    g = t.groupby(["produto", "afiliado_id", "afiliado_nome"])["peso"].sum().reset_index()
    g = g.sort_values(["produto", "afiliado_id", "peso"], ascending=[True, True, False])
    g["_k"] = list(zip(g["produto"], g["afiliado_id"]))
    canon = g.drop_duplicates("_k").set_index("_k")["afiliado_nome"].to_dict()
    t["_k"] = list(zip(t["produto"], t["afiliado_id"]))
    variantes = t.groupby("_k")["afiliado_nome"].agg(lambda x: sorted(set(x))).to_dict()
    return canon, variantes


def enriquecer(ped, ref, cb):
    """Adiciona data do pedido real, coorte, fase, tipo de agente."""
    # data do pedido, a partir da base de PEDIDOS (fonte confiavel)
    mapa_data, mapa_data_prod = {}, {}
    if len(ped):
        ped["data_pedido"] = ped["date_created"].fillna(ped["order_date"])
        p = ped.dropna(subset=["order_id", "data_pedido"])
        mapa_data = (p.groupby(p["order_id"].astype(str))["data_pedido"]
                      .min().to_dict())
        # chave mais segura: Order ID escopado por produto
        mapa_data_prod = (p.groupby([p["produto"].astype(str),
                                     p["order_id"].astype(str)])["data_pedido"]
                           .min().to_dict())
        ped["origem_data_pedido"] = "base de pedidos"

    diag = {}
    for base, col_evento in ((ref, "refund_date"), (cb, "cb_date")):
        if not len(base):
            continue
        base["data_evento"] = base[col_evento] if col_evento in base.columns else pd.NaT
        oid = base["order_id"].astype(str)
        chave_prod = list(zip(base["produto"].astype(str), oid))
        base["data_pedido"] = pd.Series(
            [mapa_data_prod.get(k) for k in chave_prod], index=base.index)
        falta_prod = base["data_pedido"].isna()
        base.loc[falta_prod, "data_pedido"] = oid[falta_prod].map(mapa_data)
        base["data_pedido"] = pd.to_datetime(base["data_pedido"], errors="coerce")
        base["origem_data_pedido"] = np.where(
            base["data_pedido"].notna(), "join com base de pedidos", "")
        falta = base["data_pedido"].isna() & base["order_date"].notna()
        base.loc[falta, "data_pedido"] = base.loc[falta, "order_date"]
        base.loc[falta, "origem_data_pedido"] = "Order Date do proprio export (fallback)"
        base["origem_data_pedido"] = base["origem_data_pedido"].replace("", "SEM DATA DE PEDIDO")

        # O export de Chargeback nao tem coluna de data do estorno: so 'Order Date'.
        # Em vez de presumir o que essa coluna e, comparamos com a data real do
        # pedido (vinda de Order Items) e deixamos o diagnostico escrito.
        if col_evento == "cb_date" and base["data_evento"].isna().all():
            comp = base.dropna(subset=["order_date"]).copy()
            comp = comp[comp["origem_data_pedido"].str.startswith("join")]
            if len(comp):
                mesmo_dia = float((comp["order_date"].dt.date ==
                                   comp["data_pedido"].dt.date).mean())
                gap = float((comp["order_date"] - comp["data_pedido"]).dt.days.median())
            else:
                mesmo_dia, gap = float("nan"), float("nan")
            if mesmo_dia == mesmo_dia and mesmo_dia >= 0.9:
                rot = ("data do PEDIDO - o export de chargeback nao traz data do estorno")
            elif mesmo_dia == mesmo_dia and gap > 0:
                rot = ("possivelmente a data do ESTORNO (mediana de %.0f dias depois "
                       "do pedido) - CONFIRMAR com a BuyGoods" % gap)
            else:
                rot = "indeterminada - sem casamento suficiente com a base de pedidos"
            base["data_evento"] = base["order_date"]
            base["base_data_evento"] = rot
            diag["cb"] = (mesmo_dia, gap, rot)

    for base in (ped, ref, cb):
        if not len(base):
            continue
        d = base["data_evento"] if "data_evento" in base.columns else base["data_pedido"]
        base["ano"] = d.dt.year
        base["mes"] = d.dt.to_period("M").astype(str)
        base["semana_ini"] = semana_inicio(d)
        base["semana"] = base["semana_ini"].map(rotulo_semana)
        base["dia"] = d.dt.date
        base["fase"] = d.map(fase_de)
        base["mes_pedido"] = base["data_pedido"].dt.to_period("M").astype(str)
        base["semana_pedido_ini"] = semana_inicio(base["data_pedido"])
        base["semana_pedido"] = base["semana_pedido_ini"].map(rotulo_semana)
        base["afiliado_nome"] = base["affiliate_name"].fillna("(sem nome)")
        base["afiliado_id"] = base["affiliate_id"].fillna("0")
        base.loc[base["afiliado_id"].isin(["0", "0.0", ""]), "afiliado_nome"] = \
            "Sem Afiliado / Pedido Direto"
        base["conta_interna"] = [eh_interna(n, i) for n, i in
                                 zip(base["afiliado_nome"], base["afiliado_id"])]

    # Category Name so existe na base de pedidos: propagar por produto,
    # senao as abas de nicho saem zeradas no reembolso e no chargeback
    if len(ped) and "categoria" in ped.columns and ped["categoria"].notna().any():
        cat = (ped.dropna(subset=["categoria"])
                  .groupby("produto")["categoria"]
                  .agg(lambda x: x.value_counts().index[0]).to_dict())
        for base in (ref, cb):
            if len(base):
                atual = base["categoria"] if "categoria" in base.columns else None
                vindo = base["produto"].map(cat)
                base["categoria"] = (vindo if atual is None
                                     else atual.fillna(vindo))

    # nome canonico por Affiliate ID (Orders tem peso maior: nome cadastrado)
    canon, variantes = nome_canonico([(ped, 3), (ref, 1), (cb, 1)])
    for base in (ped, ref, cb):
        if not len(base):
            continue
        chaves = list(zip(base["produto"], base["afiliado_id"]))
        base["afiliado_canon"] = pd.Series(
            [canon.get(k) for k in chaves], index=base.index).fillna(
            base["afiliado_nome"])
        base["variantes_nome"] = [" | ".join(variantes.get(k, [])) for k in chaves]
        # conta interna se QUALQUER variante do nome for interna
        base["conta_interna"] = pd.Series(
            [any(eh_interna(n, k[1]) for n in variantes.get(k, [])) for k in chaves],
            index=base.index) | base["conta_interna"]

    for base in (ref, cb):
        if not len(base):
            continue
        base["tipo_agente"] = base["agent"].map(classificar_agente)
        base["dias_ate_evento"] = (base["data_evento"] - base["data_pedido"]).dt.days
        base["semanas_ate_evento"] = np.floor(
            (base["semana_ini"] - base["semana_pedido_ini"]).dt.days / 7)

    return ped, ref, cb, diag


# =====================================================================
# 7. AGREGACOES
# =====================================================================

def cobertura_mensal(ped, meses):
    """Mes -> ('OK'|'PARCIAL', detalhe). Base para liberar ou nao o calculo de %."""
    out = {}
    hoje = hoje_ref()
    for m in meses:
        if not len(ped):
            out[m] = ("SEM BASE", "base de pedidos ausente")
            continue
        sub = ped[ped["mes"] == m]
        if not len(sub):
            out[m] = ("SEM BASE", "nenhum pedido nesse mes no arquivo")
            continue
        per = pd.Period(m, freq="M")
        ini, fim = per.start_time.date(), min(per.end_time.date(), hoje)
        total_dias = (fim - ini).days + 1
        dias_com = sub["dia"].nunique()
        pct = dias_com / total_dias if total_dias else 0
        if pct >= COBERTURA_MINIMA:
            out[m] = ("OK", "%d de %d dias com pedido" % (dias_com, total_dias))
        else:
            out[m] = ("PARCIAL", "so %d de %d dias com pedido (%.0f%%)" %
                      (dias_com, total_dias, pct * 100))
    return out


def _agg(base, chave, prefixo):
    if not len(base):
        return pd.DataFrame(columns=[chave])
    g = base.groupby(chave).agg(**{
        prefixo + "_qtd": ("order_id", "size"),
        prefixo + "_pedidos": ("order_id", pd.Series.nunique),
        prefixo + "_valor": ("amount", "sum"),
    }).reset_index()
    return g


def resumo_periodo(ped, ref, cb, chave, cobertura=None):
    """chave = 'mes' | 'semana' | 'produto' | 'dia' | 'fase'"""
    p = pd.DataFrame(columns=[chave])
    if len(ped):
        p = ped.groupby(chave).agg(
            Itens=("order_id", "size"),
            Pedidos=("order_id", pd.Series.nunique),
            Gross=("amount", "sum"),
        ).reset_index()
    com = pd.DataFrame(columns=[chave])
    if len(ref) and "commission_amount" in ref.columns and ref["commission_amount"].notna().any():
        com = ref.groupby(chave)["commission_amount"].sum().reset_index()
        com = com.rename(columns={"commission_amount": "Comissao devolvida"})
    r = _agg(ref, chave, "RF").rename(columns={
        "RF_qtd": "Reembolsos (linhas)", "RF_pedidos": "Pedidos reembolsados",
        "RF_valor": "Valor reembolsado"})
    c = _agg(cb, chave, "CB").rename(columns={
        "CB_qtd": "Chargebacks (linhas)", "CB_pedidos": "Pedidos com CB",
        "CB_valor": "Valor de chargeback"})
    out = p.merge(r, on=chave, how="outer").merge(c, on=chave, how="outer")
    if len(com):
        out = out.merge(com, on=chave, how="outer")
    for col in out.columns:
        if col != chave:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)
    out = out.sort_values(chave).reset_index(drop=True)
    out["Saida total"] = out.get("Valor reembolsado", 0) + out.get("Valor de chargeback", 0)
    out["Liquido"] = out.get("Gross", 0) - out["Saida total"]
    if cobertura is not None:
        out["Cobertura de pedidos"] = out[chave].map(
            lambda k: cobertura.get(k, ("?", ""))[0])
        out["Detalhe da cobertura"] = out[chave].map(
            lambda k: cobertura.get(k, ("?", ""))[1])
    return out


def cohort_semanal(ped, ref, cb, max_sem=COHORT_MAX_SEMANAS):
    """Linhas = semana da compra; colunas = W+0..W+n; valores = saida ($)."""
    if not len(ped):
        return pd.DataFrame(), pd.DataFrame()
    base_ped = ped.dropna(subset=["semana_ini"]).groupby("semana_ini").agg(
        Gross=("amount", "sum"), Pedidos=("order_id", pd.Series.nunique)
    ).reset_index()

    ev = []
    for base, rot in ((ref, "RF"), (cb, "CB")):
        if len(base):
            t = base.dropna(subset=["semana_pedido_ini", "semanas_ate_evento"]).copy()
            t = t[t["semanas_ate_evento"] >= 0]
            t["bucket"] = t["semanas_ate_evento"].clip(upper=max_sem + 1).astype(int)
            t["origem"] = rot
            ev.append(t[["semana_pedido_ini", "bucket", "amount", "origem"]])
    if not ev:
        return base_ped, pd.DataFrame()
    ev = pd.concat(ev, ignore_index=True)

    piv = ev.pivot_table(index="semana_pedido_ini", columns="bucket",
                         values="amount", aggfunc="sum", fill_value=0)
    piv.columns = ["W+%d" % c if c <= max_sem else "W+%d ou mais" % (max_sem + 1)
                   for c in piv.columns]
    piv = piv.reset_index().rename(columns={"semana_pedido_ini": "Semana da compra"})

    m = base_ped.rename(columns={"semana_ini": "Semana da compra"}).merge(
        piv, on="Semana da compra", how="left").fillna(0)
    # mes ao qual a semana pertence = mes da quinta-feira (convencao ISO),
    # para semanas que atravessam a virada de mes
    m["Mes"] = (pd.to_datetime(m["Semana da compra"]) + pd.Timedelta(days=3)
                ).dt.to_period("M").astype(str)
    m["Semana"] = pd.to_datetime(m["Semana da compra"]).map(rotulo_semana)
    cols_w = [c for c in m.columns if c.startswith("W+")]
    m["Saida total"] = m[cols_w].sum(axis=1)
    ordem = ["Mes", "Semana", "Pedidos", "Gross"] + cols_w + ["Saida total"]
    m = m.sort_values("Semana da compra")[ordem].reset_index(drop=True)

    # versao acumulada em %
    pct = m.copy()
    acum = 0
    for c in cols_w:
        acum = acum + m[c]
        pct[c] = np.where(m["Gross"] > 0, acum / m["Gross"], np.nan)
    pct["Saida total"] = np.where(m["Gross"] > 0, m["Saida total"] / m["Gross"], np.nan)
    return m, pct


def parciais_para_cb(ped, ref, cb):
    """Reembolso parcial (save-the-sale) que depois virou chargeback no mesmo pedido."""
    if not len(ref) or not len(cb) or not len(ped):
        return pd.DataFrame()
    gross = ped.dropna(subset=["order_id"]).groupby(
        ped["order_id"].astype(str))["amount"].sum()
    r = ref.dropna(subset=["order_id", "amount"]).copy()
    r["valor_pedido"] = r["order_id"].astype(str).map(gross)
    r = r[r["valor_pedido"].notna() & (r["valor_pedido"] > 0)]
    r["% do pedido"] = r["amount"] / r["valor_pedido"]
    r["Parcial"] = r["% do pedido"] < 0.9
    ids_cb = set(cb["order_id"].dropna().astype(str))
    r["Virou chargeback"] = r["order_id"].astype(str).isin(ids_cb)
    out = r[r["Parcial"] & r["Virou chargeback"]].copy()
    if not len(out):
        return pd.DataFrame()
    cols = ["produto", "order_id", "afiliado_nome", "data_evento", "amount",
            "valor_pedido", "% do pedido", "reason", "agent", "tipo"]
    out = out[[c for c in cols if c in out.columns]].rename(columns={
        "produto": "Produto", "order_id": "Order ID", "afiliado_nome": "Afiliado",
        "data_evento": "Data do reembolso", "amount": "Valor reembolsado",
        "valor_pedido": "Valor do pedido", "reason": "Motivo", "agent": "Agente",
        "tipo": "Type"})
    return out.sort_values("Data do reembolso")


def cohort_por_produto(ped, ref, cb, min_gross=50000, max_sem=COHORT_MAX_SEMANAS):
    """Mesma matriz de coorte, uma por produto, em formato longo.

    Semana com base pequena produz taxa instavel (um pedido reembolsado vira
    20%), entao o gross da semana vai junto para o leitor calibrar - e a
    coluna 'Amostra' marca as semanas fracas em vez de deixar o numero solto.
    """
    if not len(ped):
        return pd.DataFrame()
    grandes = (ped.groupby("produto")["amount"].sum()
                  .pipe(lambda s: s[s >= min_gross]).index.tolist())
    linhas = []
    for prod in sorted(grandes):
        pp = ped[ped["produto"] == prod]
        rr = ref[ref["produto"] == prod] if len(ref) else ref
        cc = cb[cb["produto"] == prod] if len(cb) else cb
        v, pc = cohort_semanal(pp, rr, cc, max_sem)
        if not len(pc):
            continue
        pc = pc.copy()
        pc.insert(0, "Produto", prod)
        pc["Amostra"] = np.where(pc["Gross"] < 5000, "baixa - taxa instavel", "ok")
        linhas.append(pc)
    return pd.concat(linhas, ignore_index=True) if linhas else pd.DataFrame()


def por_decisor(ref, cb, chave="mes"):
    linhas = []
    for base, rot in ((ref, "Reembolso"), (cb, "Chargeback")):
        if not len(base) or "tipo_agente" not in base.columns:
            continue
        g = base.groupby([chave, "tipo_agente"]).agg(
            Qtd=("order_id", "size"), Valor=("amount", "sum")).reset_index()
        g["Evento"] = rot
        linhas.append(g)
    if not linhas:
        return pd.DataFrame(), pd.DataFrame()
    d = pd.concat(linhas, ignore_index=True)
    piv_v = d[d["Evento"] == "Reembolso"].pivot_table(
        index=chave, columns="tipo_agente", values="Valor",
        aggfunc="sum", fill_value=0).reset_index()
    tot = piv_v.drop(columns=[chave]).sum(axis=1)
    piv_p = piv_v.copy()
    for c in piv_p.columns:
        if c != chave:
            piv_p[c] = np.where(tot > 0, piv_v[c] / tot, np.nan)
    piv_v["Total"] = tot
    return piv_v, piv_p


def ranking_afiliados(ped, ref, cb, internas=False):  # noqa: C901
    def prep(base, pref):
        if not len(base):
            return pd.DataFrame(columns=["afiliado_canon"])
        b = base[base["conta_interna"] == internas]
        if not len(b):
            return pd.DataFrame(columns=["afiliado_canon"])
        return b.groupby(["afiliado_canon"]).agg(**{
            pref + " qtd": ("order_id", "size"),
            pref + " valor": ("amount", "sum")}).reset_index()

    def prep_com(base):
        if not len(base) or "commission_amount" not in base.columns:
            return pd.DataFrame(columns=["afiliado_canon"])
        b = base[base["conta_interna"] == internas]
        if not len(b) or b["commission_amount"].isna().all():
            return pd.DataFrame(columns=["afiliado_canon"])
        return b.groupby(["afiliado_canon"])["commission_amount"].sum() \
                .reset_index().rename(columns={"commission_amount": "Comissao devolvida"})

    p = prep(ped, "Pedidos")
    r = prep(ref, "Reembolso")
    c = prep(cb, "Chargeback")
    if not len(p) and not len(r) and not len(c):
        return pd.DataFrame()
    out = p.merge(r, on=["afiliado_canon"], how="outer") \
           .merge(c, on=["afiliado_canon"], how="outer")
    cm = prep_com(ref)
    if len(cm):
        out = out.merge(cm, on=["afiliado_canon"], how="outer")
    for col in out.columns:
        if col not in ("afiliado_canon",):
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)
    out = out.rename(columns={"Pedidos valor": "Gross", "afiliado_canon": "Afiliado"})
    ids = pd.concat([b[["afiliado_canon", "produto", "afiliado_id"]]
                     for b in (ped, ref, cb) if len(b)], ignore_index=True)
    ids = ids.drop_duplicates()
    agg = ids.groupby("afiliado_canon").agg(
        **{"Contas (produto:ID)": ("produto", lambda x: ""),
           "Produtos": ("produto", lambda x: len(set(x)))}).reset_index()
    pares = (ids.assign(par=ids["produto"].astype(str) + ":" + ids["afiliado_id"].astype(str))
                .groupby("afiliado_canon")["par"]
                .agg(lambda x: " | ".join(sorted(set(x))[:8])).reset_index())
    agg = agg.drop(columns=["Contas (produto:ID)"]).merge(pares, on="afiliado_canon")
    agg = agg.rename(columns={"afiliado_canon": "Afiliado", "par": "Contas (produto:ID)"})
    out = out.merge(agg, on="Afiliado", how="left")
    out["Saida total"] = out.get("Reembolso valor", 0) + out.get("Chargeback valor", 0)
    out["Liquido"] = out.get("Gross", 0) - out["Saida total"]
    return out.sort_values("Gross", ascending=False).reset_index(drop=True)


# =====================================================================
# 8. EXCEL
# =====================================================================

def escrever_excel(caminho, abas, meta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.chart import LineChart, BarChart, Reference

    wb = Workbook()
    wb.remove(wb.active)
    fina = Side(style="thin", color="D9D9D9")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)

    # --- LEIA-ME
    ws = wb.create_sheet("LEIA-ME")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "TigerOffers - Analise anual de pedidos, reembolsos e chargebacks"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=NAVY)
    linha = 3
    for titulo, texto in meta:
        ws.cell(row=linha, column=1, value=titulo).font = Font(
            name="Arial", size=10, bold=True, color=BLUE)
        cel = ws.cell(row=linha, column=2, value=texto)
        cel.font = Font(name="Arial", size=10)
        cel.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[linha].height = max(15, 13 * (len(texto) // 110 + 1))
        linha += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 118

    # --- demais abas
    for item in abas:
        nome, df, fmt = item[0], item[1], item[2]
        formulas = item[3] if len(item) > 3 else {}
        if df is None or not len(df):
            continue
        col_idx = {str(c): k for k, c in enumerate(df.columns, start=1)}
        ws = wb.create_sheet(nome[:31])
        ws.sheet_view.showGridLines = False
        ws.cell(row=1, column=1, value=nome).font = Font(
            name="Arial", size=13, bold=True, color=NAVY)
        cab = 3
        for j, col in enumerate(df.columns, start=1):
            c = ws.cell(row=cab, column=j, value=str(col))
            c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=BLUE)
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = borda
        for i, (_, row) in enumerate(df.iterrows(), start=cab + 1):
            for j, col in enumerate(df.columns, start=1):
                v = row[col]
                if v is pd.NaT or (v is not None and not isinstance(v, (list, dict))
                                   and pd.isna(v) is True):
                    v = None
                elif isinstance(v, (pd.Timestamp, datetime, date)):
                    v = pd.Timestamp(v).strftime("%d/%m/%Y")
                elif isinstance(v, (np.integer,)):
                    v = int(v)
                elif isinstance(v, (np.floating,)):
                    v = float(v)
                elif isinstance(v, (np.bool_,)):
                    v = bool(v)
                c = ws.cell(row=i, column=j, value=v)
                c.font = Font(name="Arial", size=10)
                c.border = borda
                if (i - cab) % 2 == 0:
                    c.fill = PatternFill("solid", fgColor=CINZA)
                # % como formula viva quando os dois operandos estao na aba
                if str(col) in formulas and v is not None:
                    espec = formulas[str(col)]
                    if isinstance(espec, tuple):           # divisao num/den
                        num, den = espec
                        if num in col_idx and den in col_idx:
                            c.value = "=IFERROR(%s%d/%s%d,\"\")" % (
                                get_column_letter(col_idx[num]), i,
                                get_column_letter(col_idx[den]), i)
                    elif isinstance(espec, str):           # template livre
                        expr, ok = espec, True
                        for nome_ref in re.findall(r"\{([^}]+)\}", espec):
                            if nome_ref not in col_idx:
                                ok = False
                                break
                            expr = expr.replace(
                                "{%s}" % nome_ref,
                                "%s%d" % (get_column_letter(col_idx[nome_ref]), i))
                        if ok:
                            c.value = expr
                f = fmt.get(col)
                if f == "moeda":
                    c.number_format = '$#,##0.00;($#,##0.00);-'
                elif f == "pct":
                    c.number_format = '0.0%;-0.0%;-'
                elif f == "int":
                    c.number_format = '#,##0;-#,##0;-'
        ws.freeze_panes = ws.cell(row=cab + 1, column=1)
        ws.auto_filter.ref = "A%d:%s%d" % (cab, get_column_letter(len(df.columns)),
                                           cab + len(df))
        for j, col in enumerate(df.columns, start=1):
            try:
                maxlen = df[col].astype(str).str.len().max()
                maxlen = 10 if pd.isna(maxlen) else int(maxlen)
            except Exception:
                maxlen = 10
            largura = max(11, len(str(col)) + 4, min(38, maxlen + 3))
            ws.column_dimensions[get_column_letter(j)].width = min(largura, 42)
        # escala de cor nas colunas de %
        for j, col in enumerate(df.columns, start=1):
            if fmt.get(col) == "pct" and len(df) > 1:
                letra = get_column_letter(j)
                ws.conditional_formatting.add(
                    "%s%d:%s%d" % (letra, cab + 1, letra, cab + len(df)),
                    ColorScaleRule(start_type="min", start_color=VERDE.replace("#", ""),
                                   mid_type="percentile", mid_value=50,
                                   mid_color=AMARELO,
                                   end_type="max", end_color=VERMELHO))

    # --- grafico na aba Mensal, se existir
    if "Mensal" in wb.sheetnames:
        ws = wb["Mensal"]
        df = [a for a in abas if a[0] == "Mensal"][0][1]
        cols = list(df.columns)
        if "Gross" in cols and "Valor reembolsado" in cols:
            n = len(df)
            ch = LineChart()
            ch.title = "Gross x Reembolso x Chargeback por mes"
            ch.height, ch.width = 8, 22
            for nome_col in ("Gross", "Valor reembolsado", "Valor de chargeback"):
                if nome_col in cols:
                    idx = cols.index(nome_col) + 1
                    ref_ = Reference(ws, min_col=idx, min_row=3, max_row=3 + n)
                    ch.add_data(ref_, titles_from_data=True)
            cats = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
            ch.set_categories(cats)
            ws.add_chart(ch, "A%d" % (3 + n + 3))

    wb.save(caminho)
    return caminho


# =====================================================================
# 9. MAIN
# =====================================================================

DATA_REF = [None]   # ultima data observada nos arquivos; ancora todos os cortes


def hoje_ref():
    """Data de referencia = ultimo dia com dado, nao o dia em que o script roda.
    Sem isso, 'dias sem vender' muda sozinho a cada execucao e o relatorio de
    ontem deixa de bater com o de hoje."""
    return DATA_REF[0] or date.today()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--entrada", default=ENTRADA_PADRAO)
    ap.add_argument("--saida", default=SAIDA_PADRAO)
    ap.add_argument("--ano", type=int, default=ANO_PADRAO)
    args = ap.parse_args()
    os.makedirs(args.saida, exist_ok=True)

    print("=" * 78)
    print("TigerOffers - analise anual %d" % args.ano)
    print("entrada:", args.entrada)
    print("=" * 78)

    inventario, blocos = carregar(args.entrada)
    inv = pd.DataFrame(inventario)
    masters = blocos.pop("_MASTERS", {})
    ped = blocos.get("PEDIDOS", pd.DataFrame())
    ref = blocos.get("REEMBOLSOS", pd.DataFrame())
    cb = blocos.get("CHARGEBACKS", pd.DataFrame())

    if not len(ped) and not len(ref) and not len(cb):
        print("\nNada classificado. Confira o inventario acima.")
        inv.to_csv(os.path.join(args.saida, "inventario.csv"), index=False)
        return

    # Um chargeback copiado para dentro do arquivo de reembolso entra duas vezes
    # no prejuizo. Regra por CHAVE, nao por texto: linha de reembolso sem Refund
    # Date cujo Order ID ja esta na base de chargeback.
    dedup_info = None
    if len(ref) and len(cb):
        sem_rd = ref["refund_date"].isna() if "refund_date" in ref.columns else pd.Series(False, index=ref.index)
        ids_cb = set(cb["order_id"].dropna().astype(str))
        dup = sem_rd & ref["order_id"].astype(str).isin(ids_cb)
        if dup.any():
            dedup_info = (int(dup.sum()), float(ref.loc[dup, "amount"].sum()),
                          ref.loc[dup, "aba"].value_counts().to_dict())
            ref = ref[~dup].reset_index(drop=True)

    datas = []
    for base in (ped, ref, cb):
        for col in ("date_created", "order_date", "refund_date", "cb_date"):
            if len(base) and col in base.columns and base[col].notna().any():
                datas.append(base[col].max())
    if datas:
        DATA_REF[0] = max(datas).date()
        print("\n   data de referencia (ultimo dado observado):", DATA_REF[0])

    ped, ref, cb, diag = enriquecer(ped, ref, cb)
    if dedup_info:
        n, v, abas_dup = dedup_info
        print("   !! %d linhas de chargeback removidas da base de REEMBOLSO "
              "(US$ %.2f) - abas: %s" % (n, v, abas_dup))
    checks, dups = validar(ped, ref, cb, args.ano, diag)
    if dedup_info:
        n, v, abas_dup = dedup_info
        checks.loc[len(checks)] = {
            "Verificacao": "Chargebacks duplicados dentro do arquivo de reembolso",
            "Resultado": "%d linhas removidas | US$ %s" % (n, f"{v:,.2f}"),
            "Detalhe": "abas afetadas: %s. Sao linhas sem Refund Date cujo Order ID ja "
                       "consta na base de chargeback - contariam o mesmo prejuizo duas "
                       "vezes. Consequencia: se a aba ficou vazia, aquele produto esta "
                       "SEM base de reembolso." % abas_dup}

    # recorte do ano
    for nome, base in (("pedidos", ped), ("reembolsos", ref), ("chargebacks", cb)):
        if len(base):
            fora = int((base["ano"] != args.ano).sum())
            if fora:
                checks.loc[len(checks)] = {
                    "Verificacao": "Linhas fora de %d em %s" % (args.ano, nome),
                    "Resultado": "%d linhas removidas do recorte" % fora,
                    "Detalhe": "mantidas nos CSVs brutos"}
    ped_a = ped[ped["ano"] == args.ano] if len(ped) else ped
    ref_a = ref[ref["ano"] == args.ano] if len(ref) else ref
    cb_a = cb[cb["ano"] == args.ano] if len(cb) else cb

    for flag in EXCLUIR_DE_GROSS:
        if len(ped_a) and flag in ped_a.columns and ped_a[flag].notna().any():
            m = flag_true(ped_a[flag])
            if m.any():
                checks.loc[len(checks)] = {
                    "Verificacao": "Excluidos do Gross (%s)" % flag,
                    "Resultado": "%d itens | US$ %s" %
                                 (int(m.sum()), f"{ped_a.loc[m, 'amount'].sum():,.2f}"),
                    "Detalhe": "pagamento recusado nunca foi capturado; manter no Gross "
                               "inflaria o denominador e reduziria a taxa de reembolso "
                               "artificialmente"}
                ped_a = ped_a[~m]

    meses = sorted(set(list(ped_a.get("mes", pd.Series(dtype=str)).dropna()) +
                       list(ref_a.get("mes", pd.Series(dtype=str)).dropna()) +
                       list(cb_a.get("mes", pd.Series(dtype=str)).dropna())))
    cob = cobertura_mensal(ped_a, meses)

    mensal = resumo_periodo(ped_a, ref_a, cb_a, "mes", cob)
    semanal = resumo_periodo(ped_a, ref_a, cb_a, "semana")
    produtos = resumo_periodo(ped_a, ref_a, cb_a, "produto")
    fases = resumo_periodo(ped_a, ref_a, cb_a, "fase")

    # taxas: so quando a cobertura permite
    def taxas(df, tem_cobertura=True):
        df = df.copy()
        ok = (df["Cobertura de pedidos"] == "OK") if (tem_cobertura and
              "Cobertura de pedidos" in df.columns) else pd.Series(True, index=df.index)
        for nome, num in (("% Reembolso ($)", "Valor reembolsado"),
                          ("% Chargeback ($)", "Valor de chargeback"),
                          ("% Saida total ($)", "Saida total")):
            df[nome] = np.where(ok & (df.get("Gross", 0) > 0),
                                df.get(num, 0) / df.get("Gross", 1), np.nan)
        df["% Pedidos reembolsados"] = np.where(
            ok & (df.get("Pedidos", 0) > 0),
            df.get("Pedidos reembolsados", 0) / df.get("Pedidos", 1), np.nan)
        df["Alerta"] = [next(r for lim, r in ESCALA_ALERTA if (v or 0) >= lim)
                        if not pd.isna(v) else "sem base"
                        for v in df["% Saida total ($)"]]
        return df

    nota_cb = ""
    if len(cb_a) and "base_data_evento" in cb_a.columns:
        nota_cb = str(cb_a["base_data_evento"].iloc[0])
    if nota_cb:
        for _df in (mensal, semanal):
            _df["Base do chargeback"] = nota_cb

    mensal = taxas(mensal)
    semanal = taxas(semanal, tem_cobertura=False)
    produtos = taxas(produtos, tem_cobertura=False)
    fases = taxas(fases, tem_cobertura=False)

    # base coorte: reembolso/CB atribuidos ao mes do PEDIDO que os gerou
    ref_c = ref_a.copy()
    cb_c = cb_a.copy()
    for b in (ref_c, cb_c):
        if len(b):
            b["mes"] = b["mes_pedido"]
    ref_c = ref_c[ref_c["mes"].notna()] if len(ref_c) else ref_c
    cb_c = cb_c[cb_c["mes"].notna()] if len(cb_c) else cb_c
    mensal_coorte = taxas(resumo_periodo(ped_a, ref_c, cb_c, "mes", cob))
    # coorte recente ainda nao maturou: a politica interna de reembolso e de 60 dias
    hoje = pd.Timestamp(hoje_ref())
    mensal_coorte["Dias de maturacao"] = [
        int((hoje - pd.Period(m, freq="M").end_time).days) for m in mensal_coorte["mes"]]
    mensal_coorte["Maturidade"] = np.where(
        mensal_coorte["Dias de maturacao"] >= 60, "MADURA (>=60d)",
        "IMATURA - taxa vai subir")
    for c in ("% Reembolso ($)", "% Chargeback ($)", "% Saida total ($)",
              "% Pedidos reembolsados"):
        mensal_coorte.loc[mensal_coorte["Maturidade"].str.startswith("IMATURA"),
                          "Alerta"] = "coorte imatura"

    coh_v, coh_p = cohort_semanal(ped_a, ref_a, cb_a)
    coh_prod = cohort_por_produto(ped_a, ref_a, cb_a)
    dec_v, dec_p = por_decisor(ref_a, cb_a, "mes")
    dec_prod_v, _ = por_decisor(ref_a, cb_a, "produto")
    afil = ranking_afiliados(ped_a, ref_a, cb_a, internas=False)
    internas = ranking_afiliados(ped_a, ref_a, cb_a, internas=True)

    def cruzar(base, campo, rotulo):
        if not len(base) or campo not in base.columns or base[campo].isna().all():
            return pd.DataFrame()
        t = base.pivot_table(index="mes", columns=campo, values="amount",
                             aggfunc="sum", fill_value=0).reset_index()
        t.insert(1, "Recorte", rotulo)
        t["Total"] = t.drop(columns=["mes", "Recorte"]).sum(axis=1)
        return t

    tipo_rec = pd.concat([cruzar(ref_a, "tipo", "Reembolso por Type"),
                          cruzar(ref_a, "is_recurring", "Reembolso por Is recurring")],
                         ignore_index=True) if len(ref_a) else pd.DataFrame()
    parciais = parciais_para_cb(ped_a, ref_a, cb_a)

    # custos: so montam a aba se as colunas existirem de fato no export
    custos = pd.DataFrame()
    cols_custo = [c for c in ("shipping_cost", "insurance_cost", "taxes",
                              "external_fees", "provider_cost")
                  if len(ped_a) and c in ped_a.columns and ped_a[c].notna().any()]
    if cols_custo:
        custos = ped_a.groupby("mes").agg(
            **{"Gross": ("amount", "sum")},
            **{c: (c, "sum") for c in cols_custo}).reset_index()
        custos = custos.rename(columns={
            "shipping_cost": "Frete", "insurance_cost": "Seguro", "taxes": "Impostos",
            "external_fees": "Taxas externas", "provider_cost": "Custo do fornecedor"})
        cols_ren = [c for c in custos.columns if c not in ("mes", "Gross")]
        custos["Custo total"] = custos[cols_ren].sum(axis=1)
        custos = custos.merge(
            mensal[["mes", "Valor reembolsado", "Valor de chargeback"]],
            on="mes", how="left").fillna(0)
        custos["Saldo"] = (custos["Gross"] - custos["Custo total"]
                           - custos["Valor reembolsado"] - custos["Valor de chargeback"])
        preench = ped_a["provider_cost"].notna().mean() if "provider_cost" in ped_a.columns else 0
        custos["Cobertura do custo"] = "%.0f%% das linhas com Custo do fornecedor preenchido" % (preench * 100)

    categoria = pd.DataFrame()
    if len(ped_a) and "categoria" in ped_a.columns and ped_a["categoria"].notna().any():
        categoria = resumo_periodo(ped_a, ref_a, cb_a, "categoria")
        categoria = taxas(categoria, tem_cobertura=False)

    # ---- Black x White por produto
    def fase_bw(prod, d):
        alvo = VIRADA_BLACK_WHITE.get(prod)
        if not alvo or pd.isna(d):
            return "SEM DATA DE VIRADA"
        return "WHITE" if pd.Timestamp(d) >= pd.Timestamp(alvo) else "BLACK (mix 70/30)"

    bw = pd.DataFrame()
    if len(ped_a):
        for base in (ped_a, ref_a, cb_a):
            if len(base):
                base["fase_bw"] = [fase_bw(p, d) for p, d in
                                   zip(base["produto"], base["data_pedido"])]
        g = ped_a.groupby(["produto", "fase_bw"]).agg(
            Pedidos=("order_id", pd.Series.nunique), Gross=("amount", "sum"),
            Ini=("data_pedido", "min"), Fim=("data_pedido", "max")).reset_index()
        r = (ref_a.groupby(["produto", "fase_bw"])["amount"].sum().reset_index()
             .rename(columns={"amount": "Reembolso"})) if len(ref_a) else pd.DataFrame(columns=["produto", "fase_bw", "Reembolso"])
        c = (cb_a.groupby(["produto", "fase_bw"])["amount"].sum().reset_index()
             .rename(columns={"amount": "Chargeback"})) if len(cb_a) else pd.DataFrame(columns=["produto", "fase_bw", "Chargeback"])
        g = g.merge(r, on=["produto", "fase_bw"], how="left").merge(
            c, on=["produto", "fase_bw"], how="left").fillna({"Reembolso": 0, "Chargeback": 0})
        g["Dias"] = (pd.to_datetime(g["Fim"]) - pd.to_datetime(g["Ini"])).dt.days + 1
        g["Taxa"] = np.where(g["Gross"] > 0,
                             (g["Reembolso"] + g["Chargeback"]) / g["Gross"], np.nan)
        piv = g.pivot_table(index="produto", columns="fase_bw",
                            values=["Gross", "Reembolso", "Chargeback", "Taxa", "Dias", "Pedidos"])
        piv.columns = ["%s %s" % (a, b) for a, b in piv.columns]
        bw = piv.reset_index()
        col_b = [c for c in bw.columns if c.startswith("Taxa BLACK")]
        col_w = [c for c in bw.columns if c.startswith("Taxa WHITE")]
        if col_b and col_w:
            bw["Delta da taxa (WHITE - BLACK)"] = bw[col_w[0]] - bw[col_b[0]]
            bw["Veredito"] = np.where(bw["Delta da taxa (WHITE - BLACK)"] < -0.01, "MELHOROU",
                              np.where(bw["Delta da taxa (WHITE - BLACK)"] > 0.01, "PIOROU",
                                       "estavel"))
            bw.loc[bw[col_w[0]].isna() | bw[col_b[0]].isna(), "Veredito"] = "sem as duas fases"
        bw["Data da virada"] = bw["produto"].map(VIRADA_BLACK_WHITE)
        bw["Observacao"] = bw["produto"].map(OBS_VIRADA).fillna("")
        sem_data = sorted(bw.loc[bw["Data da virada"].isna(), "produto"].tolist())
        bw = bw[bw["Data da virada"].notna()].copy()
        bw = bw[[c for c in bw.columns if "SEM DATA DE VIRADA" not in c]]
        hoje_ts = pd.Timestamp(hoje_ref())
        bw["Dias de White ate hoje"] = [
            int((hoje_ts - pd.Timestamp(d)).days) for d in bw["Data da virada"]]
        bw["Maturidade da janela White"] = np.where(
            bw["Dias de White ate hoje"] >= 60, "madura (>=60d)",
            "IMATURA - reembolso ainda nao aconteceu, taxa White sai baixa por construcao")
        imat = bw["Maturidade da janela White"].str.startswith("IMATURA")
        bw.loc[imat & (bw["Veredito"] == "MELHOROU"), "Veredito"] = \
            "MELHOROU? - nao conclusivo, janela imatura"
        bw = bw.sort_values("Data da virada")

    # ---- cobertura de reembolso por produto (2 chaves)
    cobprod = pd.DataFrame()
    if len(ped_a):
        base = ped_a.groupby("produto").agg(
            Pedidos=("order_id", pd.Series.nunique), Gross=("amount", "sum")).reset_index()
        if "was_refunded" in ped_a.columns and ped_a["was_refunded"].notna().any():
            fl = ped_a[flag_true(ped_a["was_refunded"])]
            base = base.merge(fl.groupby("produto").agg(
                **{"Pedidos com flag Was Refunded": ("order_id", pd.Series.nunique),
                   "Valor desses pedidos": ("amount", "sum")}).reset_index(),
                on="produto", how="left")
        if len(ref_a):
            base = base.merge(ref_a.groupby("produto").agg(
                **{"Pedidos no arquivo de reembolso": ("order_id", pd.Series.nunique),
                   "Valor reembolsado": ("amount", "sum")}).reset_index(),
                on="produto", how="left")
        base = base.fillna(0)
        if "Pedidos com flag Was Refunded" in base.columns:
            base["Gap (flag - arquivo)"] = (base["Pedidos com flag Was Refunded"]
                                            - base.get("Pedidos no arquivo de reembolso", 0))
            base["Situacao"] = np.where(
                base.get("Pedidos no arquivo de reembolso", 0) == 0,
                "SEM BASE DE REEMBOLSO - nao calcular taxa",
                np.where(base["Gap (flag - arquivo)"].abs() /
                         base["Pedidos com flag Was Refunded"].replace(0, np.nan) > 0.1,
                         "COBERTURA PARCIAL", "ok"))
        cobprod = base.sort_values("Gross", ascending=False)

    # ---- reconciliacao contra o Master oficial
    recon = pd.DataFrame()
    if "ACCOUNTS" in masters and len(masters["ACCOUNTS"]):
        of = masters["ACCOUNTS"].copy()
        of = of[~of["__total"]]
        chave = of.columns[0]
        of["_k"] = of[chave].map(norm)
        meu = produtos.copy()
        meu["_k"] = meu["produto"].map(norm)
        chaves_of = set(of["_k"])
        meu["_k"] = [ALIAS_MASTER.get(k, k) if k not in chaves_of else k
                     for k in meu["_k"]]
        recon = of[["_k", chave, "Gross Sales", "Refunds", "Chargebacks", "Net Sales"]].merge(
            meu[["_k", "Gross", "Valor reembolsado", "Valor de chargeback"]],
            on="_k", how="outer")
        recon = recon.rename(columns={chave: "Conta (oficial)",
                                      "Gross Sales": "Gross oficial",
                                      "Refunds": "Reembolso oficial",
                                      "Chargebacks": "CB oficial",
                                      "Net Sales": "Net oficial",
                                      "Gross": "Gross calculado",
                                      "Valor reembolsado": "Reembolso calculado",
                                      "Valor de chargeback": "CB calculado"})
        for a, b, nome in (("Gross oficial", "Gross calculado", "Gap Gross"),
                           ("Reembolso oficial", "Reembolso calculado", "Gap Reembolso"),
                           ("CB oficial", "CB calculado", "Gap CB")):
            recon[nome] = (pd.to_numeric(recon[a], errors="coerce").fillna(0)
                           - pd.to_numeric(recon[b], errors="coerce").fillna(0))
        recon["% do gap de reembolso"] = np.where(
            pd.to_numeric(recon["Reembolso oficial"], errors="coerce").fillna(0) > 0,
            recon["Gap Reembolso"] / pd.to_numeric(recon["Reembolso oficial"], errors="coerce"),
            np.nan)
        recon = recon.drop(columns=["_k"]).sort_values("Gross oficial", ascending=False)

    # ---- progressao mensal dos afiliados
    prog = pd.DataFrame()
    if len(ped_a):
        ext = ped_a[~ped_a["conta_interna"]]
        if len(ext):
            g = ext.pivot_table(index="afiliado_canon", columns="mes",
                                values="amount", aggfunc="sum", fill_value=0)
            g = g.reset_index().rename(columns={"afiliado_canon": "Afiliado"})
            meses_ord = [c for c in g.columns if c != "Afiliado"]
            g["Gross total"] = g[meses_ord].sum(axis=1)
            rr = (ref_a[~ref_a["conta_interna"]].groupby("afiliado_canon")["amount"]
                  .sum().reset_index().rename(columns={"afiliado_canon": "Afiliado",
                                                       "amount": "Reembolso"})) if len(ref_a) else pd.DataFrame(columns=["Afiliado", "Reembolso"])
            cc = (cb_a[~cb_a["conta_interna"]].groupby("afiliado_canon")["amount"]
                  .sum().reset_index().rename(columns={"afiliado_canon": "Afiliado",
                                                       "amount": "Chargeback"})) if len(cb_a) else pd.DataFrame(columns=["Afiliado", "Chargeback"])
            g = g.merge(rr, on="Afiliado", how="left").merge(cc, on="Afiliado", how="left")
            g[["Reembolso", "Chargeback"]] = g[["Reembolso", "Chargeback"]].fillna(0)
            g["Taxa de saida"] = np.where(g["Gross total"] > 0,
                                          (g["Reembolso"] + g["Chargeback"]) / g["Gross total"],
                                          np.nan)
            ult = ext.groupby("afiliado_canon")["data_pedido"].max()
            pri = ext.groupby("afiliado_canon")["data_pedido"].min()
            nmes = ext.groupby("afiliado_canon")["mes"].nunique()
            g["Primeira venda"] = g["Afiliado"].map(pri)
            g["Ultima venda"] = g["Afiliado"].map(ult)
            g["Meses ativos"] = g["Afiliado"].map(nmes)
            hoje_ts = pd.Timestamp(hoje_ref())
            g["Dias sem vender"] = (hoje_ts - pd.to_datetime(g["Ultima venda"])).dt.days
            # tendencia: ultimos 3 meses com dado x 3 anteriores
            if len(meses_ord) >= 4:
                rec, ant = meses_ord[-3:], meses_ord[-6:-3]
                g["Gross 3m recentes"] = g[rec].sum(axis=1)
                g["Gross 3m anteriores"] = g[ant].sum(axis=1)
                g["Variacao 3m"] = np.where(
                    g["Gross 3m anteriores"] > 0,
                    g["Gross 3m recentes"] / g["Gross 3m anteriores"] - 1, np.nan)
            def status(r):
                if r["Dias sem vender"] > 30:
                    return "PAROU (>30d sem vender)"
                if r["Dias sem vender"] > 7:
                    return "ESFRIANDO (>7d sem vender)"
                v = r.get("Variacao 3m", np.nan)
                if v == v and v < -0.3:
                    return "CAINDO"
                if v == v and v > 0.3:
                    return "CRESCENDO"
                return "ESTAVEL"
            g["Status"] = g.apply(status, axis=1)
            g["Alerta de qualidade"] = np.where(
                g["Taxa de saida"] >= 0.30, "CRITICO - saida >=30%",
                np.where(g["Taxa de saida"] >= 0.20, "ATENCAO - saida 20-29%", ""))
            prog = g.sort_values("Gross total", ascending=False)

    # ---- produtos positivos e negativos (usa Net Sales oficial quando existir)
    posneg = pd.DataFrame()
    if len(produtos):
        posneg = produtos[["produto", "Pedidos", "Gross", "Valor reembolsado",
                           "Valor de chargeback", "Saida total",
                           "% Saida total ($)"]].copy()
        if len(recon):
            of = recon[["Conta (oficial)", "Gross oficial", "Net oficial"]].copy()
            of["_k"] = of["Conta (oficial)"].map(norm)
            posneg["_k"] = posneg["produto"].map(norm)
            kof = set(of["_k"])
            posneg["_k"] = [ALIAS_MASTER.get(k, k) if k not in kof else k
                            for k in posneg["_k"]]
            posneg = posneg.merge(of.drop(columns=["Conta (oficial)"]), on="_k",
                                  how="left").drop(columns=["_k"])
            posneg["Margem liquida oficial"] = np.where(
                pd.to_numeric(posneg["Gross oficial"], errors="coerce").fillna(0) > 0,
                pd.to_numeric(posneg["Net oficial"], errors="coerce")
                / pd.to_numeric(posneg["Gross oficial"], errors="coerce"), np.nan)
        # tendencia: ultimos 60 dias x periodo anterior
        corte60 = pd.Timestamp(hoje_ref()) - pd.Timedelta(days=60)
        if len(ped_a):
            rec = ped_a[ped_a["data_pedido"] >= corte60].groupby("produto")["amount"].sum()
            ant = ped_a[ped_a["data_pedido"] < corte60].groupby("produto")["amount"].sum()
            posneg["Gross 60d"] = posneg["produto"].map(rec).fillna(0)
            posneg["Gross antes"] = posneg["produto"].map(ant).fillna(0)
            rr60 = ref_a[ref_a["data_evento"] >= corte60].groupby("produto")["amount"].sum() if len(ref_a) else pd.Series(dtype=float)
            posneg["Reembolso 60d"] = posneg["produto"].map(rr60).fillna(0)
            posneg["Taxa 60d"] = np.where(posneg["Gross 60d"] > 0,
                                          posneg["Reembolso 60d"] / posneg["Gross 60d"], np.nan)
        if "Gross oficial" in posneg.columns:
            go = pd.to_numeric(posneg["Gross oficial"], errors="coerce")
            gc = pd.to_numeric(posneg["Gross"], errors="coerce")
            gap = (go - gc).abs() / go.replace(0, np.nan)
            posneg["Base de pedidos"] = np.where(
                gap > 0.15, "INCOMPLETA - gross calculado difere >15% do oficial",
                np.where(go.isna(), "sem conta no Master", "ok"))
            ruim = posneg["Base de pedidos"].str.startswith("INCOMPLETA")
            for c in ("% Saida total ($)", "Taxa 60d"):
                if c in posneg.columns:
                    posneg.loc[ruim, c] = np.nan

        def classificar(r):
            if str(r.get("Base de pedidos", "")).startswith("INCOMPLETA"):
                return "NAO CLASSIFICADO - base de pedidos incompleta"
            net = r.get("Net oficial", np.nan)
            if net == net and net < 0:
                return "NEGATIVO - net oficial no vermelho"
            t = r.get("% Saida total ($)", np.nan)
            if t == t and t >= 0.30:
                return "NEGATIVO - saida >=30%"
            if t == t and t >= 0.20:
                return "ATENCAO - saida 20-29%"
            m = r.get("Margem liquida oficial", np.nan)
            if m == m and m >= 0.20:
                return "POSITIVO - margem >=20%"
            return "NEUTRO"
        posneg["Classificacao"] = posneg.apply(classificar, axis=1)
        posneg = posneg.sort_values("Gross", ascending=False)

    motivos = pd.DataFrame()
    if len(ref_a) and ref_a["reason"].notna().any():
        motivos = ref_a.pivot_table(index="reason", columns="mes", values="amount",
                                    aggfunc="sum", fill_value=0).reset_index()
        motivos["Total"] = motivos.drop(columns=["reason"]).sum(axis=1)
        motivos = motivos.sort_values("Total", ascending=False)

    # helpdesk interno
    helpdesk = pd.DataFrame()
    corte = pd.Timestamp(DATA_HELPDESK_INTERNO)
    if len(ref_a):
        pos = ref_a[ref_a["data_evento"] >= corte]
        if len(pos):
            helpdesk = pos.pivot_table(index="dia", columns="tipo_agente",
                                       values="amount", aggfunc="sum",
                                       fill_value=0).reset_index()
            helpdesk["Total"] = helpdesk.drop(columns=["dia"]).sum(axis=1)

    cols_dup = ["produto", "order_id", "affiliate_id", "afiliado_nome",
                "data_evento", "amount", "reason", "agent", "arquivo", "aba"]
    if len(dups):
        dups_view = dups[[c for c in cols_dup if c in dups.columns]].rename(
            columns={"produto": "Produto", "order_id": "Order ID",
                     "affiliate_id": "Affiliate ID", "afiliado_nome": "Afiliado",
                     "data_evento": "Data do reembolso", "amount": "Valor",
                     "reason": "Motivo", "agent": "Agente",
                     "arquivo": "Arquivo", "aba": "Aba"})
    else:
        dups_view = pd.DataFrame()

    # ---- Excel
    fmt_mensal = {"Dias de maturacao": "int", "Comissao devolvida": "moeda",
                  "Gross": "moeda", "Valor reembolsado": "moeda",
                  "Valor de chargeback": "moeda", "Saida total": "moeda",
                  "Liquido": "moeda", "Pedidos": "int", "Itens": "int",
                  "Reembolsos (linhas)": "int", "Pedidos reembolsados": "int",
                  "Chargebacks (linhas)": "int", "Pedidos com CB": "int",
                  "% Reembolso ($)": "pct", "% Chargeback ($)": "pct",
                  "% Saida total ($)": "pct", "% Pedidos reembolsados": "pct"}
    fmt_coh_v = {c: ("moeda" if c.startswith("W+") or c in ("Gross", "Saida total")
                     else ("int" if c == "Pedidos" else None)) for c in coh_v.columns} if len(coh_v) else {}
    fmt_coh_p = {c: ("pct" if c.startswith("W+") or c == "Saida total"
                     else ("moeda" if c == "Gross" else ("int" if c == "Pedidos" else None)))
                 for c in coh_p.columns} if len(coh_p) else {}
    fmt_dec_v = {c: "moeda" for c in dec_v.columns if c != "mes"} if len(dec_v) else {}
    fmt_dec_p = {c: "pct" for c in dec_p.columns if c != "mes"} if len(dec_p) else {}
    fmt_af = {"Comissao devolvida": "moeda", "Gross": "moeda", "Reembolso valor": "moeda",
              "Chargeback valor": "moeda", "Saida total": "moeda",
              "Liquido": "moeda", "Pedidos qtd": "int",
              "Reembolso qtd": "int", "Chargeback qtd": "int"}

    meta = [
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Data de referencia", "%s - ultimo dia com dado nos arquivos. Todos os cortes de "
                              "'dias sem vender', maturacao e janela de 60 dias sao contados a "
                              "partir dai, nao da data em que o script rodou, para o relatorio "
                              "de hoje continuar batendo com o de ontem."
                              % hoje_ref().strftime("%d/%m/%Y")),
        ("Fonte", "arquivos lidos da pasta de entrada, sem edicao manual. "
                  "Inventario completo na aba 'Inventario'."),
        ("Recorte", "ano %d. Linhas fora do ano ficam nos CSVs brutos, nao nas abas." % args.ano),
        ("Base caixa x coorte", "'Mensal', 'Semanal', 'Produtos' e 'Fases CS' usam a DATA DO "
                                "EVENTO (quando o reembolso/chargeback aconteceu). "
                                "'Mensal Coorte' e as duas abas de Cohort usam a DATA DO PEDIDO "
                                "(o evento e atribuido ao mes/semana em que a venda foi feita). "
                                "As duas leituras respondem perguntas diferentes e nunca vao bater "
                                "linha a linha - isso e esperado, nao e erro."),
        ("Cohort por produto", "mesma matriz aberta por produto, so para os que passam de $50 mil "
                               "de gross no ano. A coluna Gross da semana fica ao lado da taxa de "
                               "proposito: semana com base pequena produz taxa instavel, e a coluna "
                               "'Amostra' marca as que estao abaixo de $5 mil."),
        ("Coorte imatura", "meses e semanas recentes ainda nao acumularam todos os reembolsos que "
                           "vao acontecer (a politica interna e de 60 dias). Na aba 'Mensal Coorte' "
                           "a coluna 'Maturidade' marca isso e o Alerta fica em 'coorte imatura' - "
                           "nao comparar esses meses com os antigos como se fossem finais."),
        ("Data do pedido", "vem da base de Pedidos via join por Order ID. O Order Date do "
                           "export de reembolso da BuyGoods e inconfiavel (costuma ser igual "
                           "ao Refund Date) e so e usado como fallback, sinalizado na coluna "
                           "'origem_data_pedido' dos CSVs."),
        ("System", "campo Agent/Agent Name igual a 'System' ou vazio = reembolso decidido pelo "
                   "automatico da BuyGoods, sem intervencao humana."),
        ("Agente Tiger", "somente %s. Qualquer outra sigla foi contada como agente da BuyGoods."
                         % ", ".join(sorted({a.title() for a in AGENTES_TIGER if " " in a}))),
        ("Fases", " | ".join("%s a partir de %s" % (r, i.strftime("%d/%m/%Y"))
                             for i, r in FASES[1:])),
        ("Helpdesk interno", "atendimento da Tiger no helpdesk da BuyGoods a partir de %s."
                             % DATA_HELPDESK_INTERNO.strftime("%d/%m/%Y")),
        ("Percentuais", "so sao calculados quando a base de pedidos cobre pelo menos %.0f%% dos "
                        "dias do mes. Mes com cobertura PARCIAL fica com a taxa em branco, de "
                        "proposito - denominador incompleto produz numero falso."
                        % (COBERTURA_MINIMA * 100)),
        ("Chargeback: sem data do estorno", "o export de Chargeback nao tem coluna de data do "
                                             "estorno, nem Agent Name, nem Reason. A unica data e "
                                             "'Order Date'. O script compara essa coluna com a data "
                                             "real do pedido e escreve o diagnostico na aba de "
                                             "validacao. Enquanto isso, a coluna 'Base do chargeback' "
                                             "diz em cada linha o que aquele numero representa - nao "
                                             "ler o CB mensal como caixa sem checar isso."),
        ("Decisor", "as abas de decisor cobrem SO reembolso. Chargeback e decidido pelo banco e o "
                    "export nem traz Agent Name."),
        ("Motivo do chargeback", "nao existe no export atual. Para abrir CB por motivo e preciso "
                                 "outro relatorio da BuyGoods."),
        ("Was Declined", "pedido recusado nunca teve captura: excluido do Gross. Cancelados e voids "
                         "continuam no Gross e aparecem quantificados na aba de validacao - trocar "
                         "isso e editar EXCLUIR_DE_GROSS no topo do script."),
        ("Duas chaves", "o universo de reembolso e de chargeback e conferido por duas chaves "
                        "independentes: as flags Was Refunded / Was Chargeback da base de pedidos e "
                        "os proprios arquivos de reembolso e chargeback. O gap entre as duas esta "
                        "escrito na aba de validacao."),
        ("Black x White", "virada por produto conforme a planilha oficial. Antes da data o "
                          "produto rodava MIX 70/30, nao 100% Black. BreathEaseX passou a 50/50. "
                          "Produto sem data configurada aparece como 'sem data de virada' e fica "
                          "fora da comparacao. ATENCAO: a janela White e curta e o reembolso ainda "
                          "nao maturou (politica de 60 dias), entao a taxa White sai "
                          "estruturalmente mais baixa - comparar com essa ressalva."),
        ("Reconciliacao vs Master", "compara o que este relatorio calculou com o Master Accounts "
                                    "oficial da BuyGoods, conta por conta. Gap grande = export de "
                                    "detalhe incompleto para aquele produto."),
        ("Type / Is recurring", "os valores reais dessas colunas nao foram presumidos: a aba "
                                "'Cobertura & Validacao' lista o universo encontrado no arquivo, "
                                "e a aba 'Tipo e Recorrencia' abre o valor por mes."),
        ("Commission Amount", "e a comissao devolvida pelo afiliado, NAO o valor reembolsado ao "
                              "cliente. Fica em coluna propria e nunca entra no total de reembolso."),
        ("Void Date", "linha com Void Date preenchido e cancelamento antes da captura, nao "
                      "reembolso de venda liquidada. Contada nos totais e reportada a parte na "
                      "aba de validacao."),
        ("Parciais que viraram CB", "reembolso abaixo de 90% do valor do pedido (save-the-sale) "
                                    "cujo Order ID depois aparece na base de chargeback."),
        ("Contas internas", "HelpGrid, MaxWeb/MWE, Tiger Offers LTDA, gestor five, "
                            "gestor1/gestor3.buygoods e Jose Moraes ficam fora do ranking de "
                            "afiliados (aba propria), mas continuam nos totais gerais."),
        ("Affiliate ID 0/vazio", "tratado como 'Sem Afiliado / Pedido Direto' nas tres bases."),
        ("Limitacao conhecida", "aff_id e escopado por conta BuyGoods: o mesmo numero pode ser "
                                "pessoas diferentes em contas diferentes. Colisoes listadas na "
                                "aba 'Cobertura & Validacao'."),
        ("Nome do afiliado", "reembolso e chargeback trazem o USERNAME, pedidos trazem o NOME "
                             "CADASTRADO. O agrupamento e feito por Affiliate ID e o nome exibido "
                             "e o canonico (prioriza o da base de pedidos). A coluna 'Variantes de "
                             "nome' mostra todas as grafias encontradas para o mesmo ID."),
    ]

    conf = conferencia_fontes([("PEDIDOS", ped), ("REEMBOLSOS", ref),
                               ("CHARGEBACKS", cb)], args.ano)
    fmt_conf = {c: ("moeda" if c.startswith("Valor") else
                    ("int" if c.startswith("Linhas") else None))
                for c in conf.columns} if len(conf) else {}

    FORM_TAXA = {
        "% Reembolso ($)": ("Valor reembolsado", "Gross"),
        "% Chargeback ($)": ("Valor de chargeback", "Gross"),
        "% Saida total ($)": ("Saida total", "Gross"),
        "% Pedidos reembolsados": ("Pedidos reembolsados", "Pedidos"),
    }

    abas = [
        ("Inventario", inv, {"Linhas": "int"}),
        ("Cobertura & Validacao", checks, {}),
        ("Conferencia fonte x recorte", conf, fmt_conf),
        ("Mensal", mensal, fmt_mensal, FORM_TAXA),
        ("Mensal Coorte", mensal_coorte, fmt_mensal, FORM_TAXA),
        ("Semanal", semanal, fmt_mensal, FORM_TAXA),
        ("Produtos", produtos, fmt_mensal, FORM_TAXA),
        ("Fases CS", fases, fmt_mensal, FORM_TAXA),
        ("Cohort Semanal $", coh_v, fmt_coh_v),
        ("Cohort Semanal %", coh_p, fmt_coh_p),
        ("Cohort Produto %", coh_prod,
         {c: ("pct" if c.startswith("W+") or c == "Saida total"
              else ("moeda" if c == "Gross" else ("int" if c == "Pedidos" else None)))
          for c in coh_prod.columns} if len(coh_prod) else {}),
        ("Decisor x Mes $", dec_v, fmt_dec_v),
        ("Decisor x Mes %", dec_p, fmt_dec_p),
        ("Decisor x Produto $", dec_prod_v,
         {c: "moeda" for c in dec_prod_v.columns if c != "produto"} if len(dec_prod_v) else {}),
        ("Helpdesk 17-08 em diante", helpdesk,
         {c: "moeda" for c in helpdesk.columns if c != "dia"} if len(helpdesk) else {}),
        ("Afiliados", afil, fmt_af),
        ("Contas Internas", internas, fmt_af),
        ("Produtos positivos e negativos", posneg,
         {c: ("pct" if c.startswith("%") or c.startswith("Taxa") or c.startswith("Margem")
              else ("int" if c == "Pedidos" else ("moeda" if c not in ("produto", "Classificacao") else None)))
          for c in posneg.columns} if len(posneg) else {}),
        ("Afiliados progressao", prog,
         {c: ("pct" if c.startswith("Taxa") or c.startswith("Variacao")
              else ("int" if c in ("Meses ativos", "Dias sem vender")
                    else ("moeda" if c.startswith("20") or c.startswith("Gross")
                          or c in ("Reembolso", "Chargeback") else None)))
          for c in prog.columns} if len(prog) else {}),
        ("Reconciliacao vs Master", recon,
         {c: ("pct" if c.startswith("%") else "moeda") for c in recon.columns
          if c != "Conta (oficial)"} if len(recon) else {}),
        ("Cobertura por produto", cobprod,
         {c: ("moeda" if "Valor" in c or c == "Gross" else "int")
          for c in cobprod.columns if c not in ("produto", "Situacao")} if len(cobprod) else {}),
        ("Black x White", bw.assign(**{"Produtos sem data de virada":
                                        (", ".join(sem_data) if len(bw) else "")})
         if len(bw) else bw,
         {c: ("pct" if c.startswith("Taxa") or c.startswith("Delta")
              else ("moeda" if c.split(" ")[0] in ("Gross", "Reembolso", "Chargeback")
                    else ("int" if c.split(" ")[0] in ("Dias", "Pedidos") else None)))
          for c in bw.columns} if len(bw) else {}),
        ("Custos e Saldo", custos,
         {c: "moeda" for c in custos.columns
          if c not in ("mes", "Cobertura do custo")} if len(custos) else {},
         {"Saldo": "={Gross}-{Custo total}-{Valor reembolsado}-{Valor de chargeback}",
          "Custo total": "=SUM({Frete}:{Custo do fornecedor})"}),
        ("Categoria (nicho)", categoria, fmt_mensal, FORM_TAXA),
        ("Tipo e Recorrencia", tipo_rec,
         {c: "moeda" for c in tipo_rec.columns
          if c not in ("mes", "Recorte")} if len(tipo_rec) else {}),
        ("Parciais que viraram CB", parciais,
         {"Valor reembolsado": "moeda", "Valor do pedido": "moeda",
          "% do pedido": "pct"}),
        ("Motivos", motivos, {c: "moeda" for c in motivos.columns
                              if c != "reason"} if len(motivos) else {}),
        ("Duplicidades", dups_view, {"Valor": "moeda"}),
    ]

    nome_saida = os.path.join(
        args.saida, "Analise_Ano_%d_TigerOffers_%s.xlsx" %
        (args.ano, datetime.now().strftime("%Y-%m-%d")))
    escrever_excel(nome_saida, abas, meta)

    # CSVs das bases consolidadas
    for nome, base in (("base_pedidos", ped), ("base_reembolsos", ref),
                       ("base_chargebacks", cb)):
        if len(base):
            base.to_csv(os.path.join(args.saida, nome + ".csv"), index=False)

    # ---- digest para colar
    linhas = ["=== RESUMO DO ANO %d ===" % args.ano, ""]
    tot_g = ped_a["amount"].sum() if len(ped_a) else 0
    tot_r = ref_a["amount"].sum() if len(ref_a) else 0
    tot_c = cb_a["amount"].sum() if len(cb_a) else 0
    linhas.append("Pedidos\t%d" % (ped_a["order_id"].nunique() if len(ped_a) else 0))
    linhas.append("Gross\t%.2f" % tot_g)
    linhas.append("Reembolso\t%.2f" % tot_r)
    linhas.append("Chargeback\t%.2f" % tot_c)
    if tot_g:
        linhas.append("Saida / Gross\t%.4f" % ((tot_r + tot_c) / tot_g))
    linhas.append("")
    linhas.append("--- MENSAL (colar no Sheets) ---")
    cols_dig = [c for c in ["mes", "Pedidos", "Gross", "Valor reembolsado",
                            "Valor de chargeback", "% Saida total ($)",
                            "Cobertura de pedidos"] if c in mensal.columns]
    linhas.append("\t".join(cols_dig))
    for _, r in mensal.iterrows():
        linhas.append("\t".join(
            ("%.4f" % r[c]) if isinstance(r[c], float) and not np.isnan(r[c])
            else ("" if (isinstance(r[c], float) and np.isnan(r[c])) else str(r[c]))
            for c in cols_dig))
    if len(dec_p):
        linhas.append("")
        linhas.append("--- QUEM DECIDIU O REEMBOLSO (%% do valor) ---")
        linhas.append("\t".join(str(c) for c in dec_p.columns))
        for _, r in dec_p.iterrows():
            linhas.append("\t".join(
                ("%.4f" % r[c]) if isinstance(r[c], float) and not np.isnan(r[c])
                else str(r[c]) for c in dec_p.columns))
    texto = "\n".join(linhas)
    with open(os.path.join(args.saida, "resumo_colar.txt"), "w",
              encoding="utf-8") as f:
        f.write(texto)

    print("\n" + "=" * 78)
    print(texto)
    print("=" * 78)
    print("\nArquivo gerado:", nome_saida)


if __name__ == "__main__":
    main()
