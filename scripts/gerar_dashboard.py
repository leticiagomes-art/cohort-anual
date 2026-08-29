#!/usr/bin/env python3
"""
gerar_dashboard.py
------------------
Lê os CSVs acumulados (data/base_*.csv) + AOV dos relatórios de afiliados,
monta o payload JSON e injeta no template do dashboard.

Saída: public/index.html

Uso:
    python scripts/gerar_dashboard.py

Dependências: pandas, openpyxl, numpy
"""
import sys, os, json, glob, re, unicodedata
from pathlib import Path
from datetime import datetime, timedelta, timezone
import pandas as pd
import numpy as np

ROOT         = Path(__file__).resolve().parent.parent
DATA_DIR     = ROOT / "data"
PUBLIC_DIR   = ROOT          # GitHub Pages serve a partir da raiz
AOV_DIR      = DATA_DIR / "aov"
TEMPLATE     = ROOT / "scripts" / "dashboard_template.html"
OUT          = ROOT / "index.html"

# ── helpers ───────────────────────────────────────────────────────────────────
def norm(s):
    t = unicodedata.normalize('NFKD', str(s).lower()).encode('ascii', 'ignore').decode()
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', '', t)).strip()

def sf(v):
    """safe float — NaN/None → None"""
    try:
        f = float(v)
        return None if (f != f) else round(f, 6)
    except (TypeError, ValueError):
        return None

INTERNOS = {'helpgrid', 'maxweb', 'mwe', 'tiger offers ltda',
            'gestor', 'gestor1', 'gestor3', 'jose moraes'}
NAO_AFIL  = 'Sem Afiliado / Pedido Direto'

def e_interno(nome):
    n = norm(str(nome))
    return any(i in n for i in INTERNOS)

BW = {
    'BreathEaseX': '2026-08-12', 'NervoLyn': '2026-07-24',
    'Prostafense':  '2026-07-27', 'AudiLeaf': '2026-07-28',
    'VisiumPro':    '2026-07-28', 'MaroBrain': '2026-07-28',
    'GlucoRecover': '2026-07-31', 'FloraNew':  '2026-07-31',
    'AlphaErec':    '2026-08-04', 'BoosterXT': '2026-08-07',
    'NailsCleanPro':'2026-08-12', 'LipoBliss': '2026-08-13',
    'LipoPeak':     '2026-08-14', 'MounjaMelt':'2026-08-14',
    'ReduTide':     '2026-07-07', 'RedurBurn': '2026-07-07',
    'LipoVive':     '2026-08-14', 'VigorLong': '2026-08-07',
    'Lipotrine':    '2026-08-14',
}

# ── 1. carregar CSVs ──────────────────────────────────────────────────────────
print("→ Carregando CSVs particionados...")

def carregar_particoes(subpasta):
    """
    Lê e concatena todas as partições mensais de uma base.

    Os dados ficam em data/{subpasta}/AAAA-MM.csv.gz. Se a pasta não
    existir, cai para o arquivo único antigo (compatibilidade).
    """
    pasta = DATA_DIR / subpasta
    if pasta.is_dir():
        arquivos = sorted(pasta.glob("*.csv.gz"))
        if arquivos:
            partes = [pd.read_csv(a, low_memory=False) for a in arquivos]
            df = pd.concat(partes, ignore_index=True)
            print(f"   {subpasta}: {len(arquivos)} mês(es), {len(df):,} linhas")
            return df
    # fallback: arquivo único
    legado = DATA_DIR / f"base_{subpasta}.csv.gz"
    if legado.exists():
        df = pd.read_csv(legado, low_memory=False)
        print(f"   {subpasta}: arquivo único (legado), {len(df):,} linhas")
        return df
    print(f"   {subpasta}: NENHUM DADO ENCONTRADO")
    return pd.DataFrame()

ped = carregar_particoes("pedidos")
ref = carregar_particoes("reembolsos")
cb  = carregar_particoes("chargebacks")

if not len(ped):
    print("ERRO: base de pedidos vazia — nada a gerar")
    sys.exit(1)

for d in (ped, ref, cb):
    # format='mixed' e obrigatorio aqui: o CSV acumulado mistura datas
    # "AAAA-MM-DD" (sem hora) com "AAAA-MM-DD HH:MM:SS" (com hora) na mesma
    # coluna, dependendo de qual arquivo/merge escreveu cada linha. Sem
    # format='mixed', pd.to_datetime trava no formato do PRIMEIRO valor nao
    # nulo e vira NaT em todo o resto — foi o que fez o total de reembolso
    # do ano cair de ~US$ 1,4 mi para US$ 2 mil no dashboard.
    d['data_pedido'] = pd.to_datetime(d['data_pedido'], errors='coerce', format='mixed')
    if 'data_evento' in d.columns:
        d['data_evento'] = pd.to_datetime(d['data_evento'], errors='coerce', format='mixed')

# âncora: última data com dado nos arquivos (não data de hoje)
DATA_REF = ped['data_pedido'].max().normalize()
ANO      = DATA_REF.year
print(f"   data de referência: {DATA_REF.date()} | ano: {ANO}")

ped_a = ped[ped['data_pedido'].dt.year == ANO].copy()
ref_a = ref[ref['data_pedido'].dt.year == ANO].copy()
cb_a  = cb[cb['data_pedido'].dt.year  == ANO].copy()

# semana_ini já está nos CSVs — garantir dtype
for d in (ped_a, ref_a, cb_a):
    d['semana_ini'] = pd.to_datetime(d['semana_ini'], errors='coerce')

# ── 2. AOV dos relatórios de afiliados ───────────────────────────────────────
print("→ Carregando AOV...")

def load_aov(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    cols = rows[0]
    out  = {}
    for row in rows[2:]:
        if not row[0] or str(row[0]).upper().startswith('TOTAL'):
            continue
        d = dict(zip(cols, row))
        out[norm(d['Afiliado'])] = sf(d.get('AOV Líquido'))
    return out

def latest(prefix):
    files = sorted(glob.glob(str(AOV_DIR / f"{prefix}*.xlsx")))
    return files[-1] if files else None

white_f = latest("white_") or latest("relatorio_afiliados_allProducts_2026-07-29")
black_f = latest("black_") or latest("relatorio_afiliados_allProducts_2026-06-01")

aov_w = load_aov(white_f) if white_f else {}
aov_b = load_aov(black_f) if black_f else {}
print(f"   white: {Path(white_f).name if white_f else 'n/a'} ({len(aov_w)})")
print(f"   black: {Path(black_f).name if black_f else 'n/a'} ({len(aov_b)})")

# ── 3. mensal ─────────────────────────────────────────────────────────────────
print("→ Mensal...")
meses = sorted(ped_a['mes'].dropna().unique())
mensal = []
for m in meses:
    pp = ped_a[ped_a['mes'] == m]
    rr = ref_a[ref_a['mes'] == m]
    cc = cb_a[cb_a['mes']  == m]
    g  = float(pp['amount'].sum())
    rv = float(rr['amount'].sum())
    cv = float(cc['amount'].sum())
    mensal.append({
        'mes': m, 'Pedidos': int(pp['order_id'].nunique()),
        'Gross': round(g, 2), 'Valor reembolsado': round(rv, 2),
        'Valor de chargeback': round(cv, 2),
        '% Reembolso ($)':    sf(rv / g) if g else None,
        '% Chargeback ($)':   sf(cv / g) if g else None,
        '% Saida total ($)':  sf((rv + cv) / g) if g else None,
        'Alerta': ('VERMELHO' if (rv+cv)/g >= .25 else
                   'LARANJA'  if (rv+cv)/g >= .18 else 'VERDE') if g else None,
    })

# ── 4. afiliados ──────────────────────────────────────────────────────────────
print("→ Afiliados...")
ped_ext = ped_a[
    ~ped_a['conta_interna'].astype(str).str.lower().isin(['true', '1']) &
    (ped_a['afiliado_canon'] != NAO_AFIL) &
    ~ped_a['afiliado_canon'].map(e_interno)
]

pg = ped_ext.groupby('afiliado_canon').agg(
    pedidos   = ('order_id',    'nunique'),
    gross     = ('amount',      'sum'),
    ultimo    = ('data_pedido', 'max'),
    primeiro  = ('data_pedido', 'min'),
).reset_index()

rg = ref_a[~ref_a['conta_interna'].astype(str).str.lower().isin(['true','1'])]\
         .groupby('afiliado_canon')['amount'].sum().reset_index()\
         .rename(columns={'amount': 'refund'})
cg = cb_a[~cb_a['conta_interna'].astype(str).str.lower().isin(['true','1'])]\
         .groupby('afiliado_canon')['amount'].sum().reset_index()\
         .rename(columns={'amount': 'cb'})

afs = pg.merge(rg, on='afiliado_canon', how='left')\
        .merge(cg, on='afiliado_canon', how='left')\
        .fillna({'refund': 0, 'cb': 0})

afs['saida'] = afs['refund'] + afs['cb']
afs['taxa']  = np.where(afs['gross'] > 0, afs['saida'] / afs['gross'], np.nan)
afs['dias']  = (DATA_REF - afs['ultimo']).dt.days.fillna(999).astype(int)

# variação 3 meses: gross dos últimos 3 meses vs 3 meses anteriores
meses_sorted = sorted(meses)
m_rec  = meses_sorted[-3:] if len(meses_sorted) >= 3 else meses_sorted
m_ant  = meses_sorted[-6:-3] if len(meses_sorted) >= 6 else []

gross_rec = ped_ext[ped_ext['mes'].isin(m_rec)].groupby('afiliado_canon')['amount'].sum()
gross_ant = ped_ext[ped_ext['mes'].isin(m_ant)].groupby('afiliado_canon')['amount'].sum() if m_ant else pd.Series(dtype=float)
afs['var3m'] = afs['afiliado_canon'].map(
    lambda c: sf((gross_rec.get(c, 0) - gross_ant.get(c, 1)) / gross_ant.get(c, 1))
    if gross_ant.get(c, 0) > 0 else None
)

# status
def status_fn(row):
    if row['dias'] > 30:   return 'PAROU'
    if row['dias'] > 7:    return 'ESFRIANDO'
    if row['taxa'] >= .30: return 'CRESCENDO' if (row['var3m'] or 0) > 0 else 'CAINDO'
    return 'ESTAVEL'
afs['Status'] = afs.apply(status_fn, axis=1)
afs['Alerta de qualidade'] = np.where(afs['taxa'] >= .30, 'CRITICO', '')

# gross mensal por afiliado
mes_gross = {
    m: ped_ext[ped_ext['mes'] == m].groupby('afiliado_canon')['amount'].sum()
    for m in meses
}

af_rows = []
for _, r in afs.iterrows():
    nome = r['afiliado_canon']
    n    = norm(nome)
    row  = {
        'Afiliado':             nome,
        'Gross total':          round(float(r['gross']), 2),
        'Reembolso':            round(float(r['refund']), 2),
        'Chargeback':           round(float(r['cb']),    2),
        'Taxa de saida':        sf(r['taxa']),
        'Dias sem vender':      int(r['dias']),
        'Variacao 3m':          r['var3m'],
        'Status':               r['Status'],
        'Alerta de qualidade':  r['Alerta de qualidade'],
        'aov_white':            aov_w.get(n),
        'aov_black':            aov_b.get(n),
    }
    for m in meses:
        row[m] = sf(mes_gross[m].get(nome))
    af_rows.append(row)

af_rows.sort(key=lambda r: -(r['Gross total'] or 0))

# ── 5. produtos ───────────────────────────────────────────────────────────────
print("→ Produtos...")
prods = list(ped_a.groupby('produto')['amount'].sum().sort_values(ascending=False).index)

def prod_row(p):
    pp = ped_a[ped_a['produto'] == p]
    g  = float(pp['amount'].sum())
    rv = float(ref_a[ref_a['produto'] == p]['amount'].sum())
    cv = float(cb_a[cb_a['produto']  == p]['amount'].sum())
    fim = pp['data_pedido'].max()
    return {
        'produto':              p,
        'Gross':                round(g, 2),
        'Pedidos':              int(pp['order_id'].nunique()),
        'Valor reembolsado':    round(rv, 2),
        'Valor de chargeback':  round(cv, 2),
        '% Saida total ($)':   sf((rv + cv) / g) if g else None,
        'Classificacao':        ('NEGATIVO' if g > 0 and (rv+cv)/g >= .30 else
                                 'ATENCAO'  if g > 0 and (rv+cv)/g >= .20 else
                                 'POSITIVO' if g > 0 else 'NAO CLASSIFICADO'),
        'Net oficial':          None,   # vem do Master Accounts — não calculado aqui
        'Margem liquida oficial': None,
        'Dias parado':          int((DATA_REF - fim).days) if pd.notna(fim) else None,
    }

produtos = [prod_row(p) for p in prods]

# ── 6. motivos ────────────────────────────────────────────────────────────────
mo = ref_a.groupby('reason')['amount'].sum().reset_index().rename(columns={'amount': 'Total'})
motivos = [{'reason': r['reason'], 'Total': round(float(r['Total']), 2)}
           for _, r in mo.sort_values('Total', ascending=False).head(20).iterrows()
           if pd.notna(r['reason'])]

# ── 7. cohort ─────────────────────────────────────────────────────────────────
print("→ Cohort...")

def mk_cohort(ped_b, ev_b, max_w=12):
    rows = []
    prod_sem = ped_b.groupby('semana_ini').apply(
        lambda g: g['produto'].mode().iloc[0] if len(g) else ''
    ).to_dict()
    for sem in sorted(ped_b['semana_ini'].dropna().unique()):
        pp = ped_b[ped_b['semana_ini'] == sem]
        n  = int(pp['order_id'].nunique())
        g  = float(pp['amount'].sum())
        if g < 1: continue
        d_ = int((DATA_REF - sem).days // 7)
        ev_sem = ev_b[ev_b['semana_ini'] == sem]
        ws = {}
        for w in range(1, max_w + 1):
            wf  = sem + pd.Timedelta(weeks=w)
            cnt = float(ev_sem.loc[ev_sem['data_evento'] < wf, 'amount'].sum()) \
                  if 'data_evento' in ev_sem.columns else 0
            ws[f'W{w}'] = round(cnt / g, 4) if g > 0 and w <= d_ + 1 else None
        prod = prod_sem.get(sem, '')
        virada = BW.get(prod)
        bw = 'white' if (virada and sem >= pd.Timestamp(virada)) else 'black'
        rows.append({
            's':  str(sem)[:10],
            'sf': pd.Timestamp(sem).strftime('%b %d, %Y'),
            'n': n, 'g': round(g, 2), 'd': d_, 'bw': bw, **ws
        })
    return rows

coh_rf = mk_cohort(ped_a, ref_a)
coh_cb = mk_cohort(ped_a, cb_a)

# cohort por produto (todos os 31)
prf, pcb = {}, {}
for p in prods:
    prf[p] = mk_cohort(ped_a[ped_a['produto'] == p], ref_a[ref_a['produto'] == p])
    pcb[p] = mk_cohort(ped_a[ped_a['produto'] == p], cb_a[cb_a['produto']  == p])
    for row in prf[p] + pcb[p]:
        virada = BW.get(p)
        row['bw'] = 'white' if (virada and pd.Timestamp(row['s']) >= pd.Timestamp(virada)) else 'black'

# curvas de velocidade
def curva(rows, m0, m1):
    sel = [r for r in rows if m0 <= r['s'] <= m1 and r['g'] > 0]
    if not sel: return [0] * 13
    gt = sum(r['g'] for r in sel)
    return [0] + [round(sum((r.get(f'W{i+1}') or 0) * r['g'] for r in sel) / gt, 4)
                  for i in range(12)]

yr = str(ANO)
curva_a = curva(coh_rf, f'{yr}-01-01', f'{yr}-05-31')
curva_b = curva(coh_rf, f'{yr}-06-01', f'{yr}-07-15')

# cohortResumo (W+0/W+4/W+8 ponderado, semanas maduras ≥ 8)
cohort_resumo = {}
for p in prods:
    rows_p = prf[p]
    n = len(rows_p)
    mad = [r for i, r in enumerate(rows_p) if (n - 1 - i) >= 8 and (r['g'] or 0) > 0]
    if not mad: continue
    tot = sum(r['g'] for r in mad)
    cohort_resumo[p] = {
        'w0':     round(sum((r.get('W+0') or 0) * r['g'] for r in mad) / tot, 5),
        'w4':     round(sum((r.get('W4')  or 0) * r['g'] for r in mad) / tot, 5),
        'w8':     round(sum((r.get('W8')  or 0) * r['g'] for r in mad) / tot, 5),
        'gross':  round(tot, 2),
        'semanas': len(mad),
    }

# ── 8. decisor × mês ─────────────────────────────────────────────────────────
print("→ Decisor...")
dec_p, dec_v = [], []
for m in meses:
    rr = ref_a[ref_a['mes'] == m]
    tot = float(rr['amount'].sum())
    if tot == 0: continue
    row_p = {'mes': m}; row_v = {'mes': m}
    for agente in ['System (automatico BuyGoods)', 'Agente BuyGoods',
                   'Agente Tiger (CS interno)']:
        v = float(rr[rr['tipo_agente'] == agente]['amount'].sum())
        key = agente
        row_p[key] = sf(v / tot)
        row_v[key] = round(v, 2)
    dec_p.append(row_p); dec_v.append(row_v)

# System por semana (para o gráfico de validação pós-17/08)
ref_a['dia'] = ref_a['data_evento'].dt.normalize()
ref_a['sys'] = ref_a['tipo_agente'].eq('System (automatico BuyGoods)')
sw2 = ref_a[ref_a['dia'] >= f'{yr}-06-01'].copy()
sem_col = sw2['dia'] - pd.to_timedelta(sw2['dia'].dt.dayofweek, unit='D')
wg = sw2.groupby(sem_col).apply(
    lambda x: pd.Series({'sys': x.loc[x['sys'], 'amount'].sum() / x['amount'].sum(),
                          'nn': len(x)})
)
sys_semana = [{'semana': i.strftime('%d/%m'), 'sys': round(float(v['sys']), 4),
               'n': int(v['nn'])} for i, v in wg.iterrows()]

bg = ref_a[(ref_a['dia'] >= f'{yr}-06-01') & (ref_a['dia'] <= '2026-08-16')].copy()
sem2 = bg['dia'] - pd.to_timedelta(bg['dia'].dt.dayofweek, unit='D')
bg2  = bg.groupby(sem2).apply(lambda x: x.loc[x['sys'], 'amount'].sum() / x['amount'].sum())
sys_ref = {'media': round(float(bg2.mean()), 4), 'dp': round(float(bg2.std()), 4),
           'min': round(float(bg2.min()), 4),    'max': round(float(bg2.max()), 4)}

# ── 9. salvamento pós-17/08 ───────────────────────────────────────────────────
pos = ref_a[(ref_a['dia'] >= '2026-08-17') & (ref_a['dia'] <= str(DATA_REF.date()))]
pos17 = {
    'total':         round(float(pos['amount'].sum()), 2),
    'tiger':         round(float(pos.loc[pos['tipo_agente'] == 'Agente Tiger (CS interno)', 'amount'].sum()), 2),
    'sys':           sf(pos.loc[pos['sys'], 'amount'].sum() / pos['amount'].sum()),
    'linhasHumano':  int((~pos['sys']).sum()),
    'linhasTiger':   int((pos['tipo_agente'] == 'Agente Tiger (CS interno)').sum()),
}
salvamento = {}
for tipo, x in pos.groupby('tipo_agente'):
    salv = x['reason'].astype(str).str.contains('Partial Refund - Saved', case=False, na=False)
    salvamento[tipo] = {
        'n':          len(x),
        'valor':      round(float(x['amount'].sum()), 2),
        'salvos':     int(salv.sum()),
        'taxaSalvo':  sf(salv.mean()),
        'retido':     round(float((x.loc[salv, 'valor_pedido'] - x.loc[salv, 'amount']).sum()), 2)
                      if 'valor_pedido' in x.columns else 0,
        'ticket':     round(float(x['amount'].mean()), 2),
    }

# ── 10. maxwell ───────────────────────────────────────────────────────────────
import re as re_mod, unicodedata as ud
def nn(s):
    t = ud.normalize('NFKD', str(s).lower()).encode('ascii','ignore').decode()
    return re_mod.sub(r'[^a-z0-9]', '', t)
ped_a['mw'] = ped_a['afiliado_canon'].fillna('').map(nn).str.contains('maxweb|mwe')
ref_a['mw'] = ref_a['afiliado_canon'].fillna('').map(nn).str.contains('maxweb|mwe')

mwg = ped_a[ped_a['mw']].groupby('mes')['amount'].sum()
ggo = ped_a[~ped_a['mw']].groupby('mes')['amount'].sum()
rmw = ref_a[ref_a['mw']].groupby('mes')['amount'].sum()
rsmw= ref_a[ref_a['mw'] & ref_a['mes_pedido'].notna()].groupby('mes_pedido')['amount'].sum()
rso = ref_a[(~ref_a['mw']) & ref_a['mes_pedido'].notna()].groupby('mes_pedido')['amount'].sum()
maxweb = []
for m in sorted(mwg.index):
    maxweb.append({
        'mes':         m,
        'gross':       round(float(mwg.get(m, 0)), 2),
        'refund':      round(float(rmw.get(m, 0)), 2),
        'safra':       sf(rsmw.get(m, 0) / mwg.get(m, 1)) if mwg.get(m, 0) else None,
        'safra_resto': sf(rso.get(m, 0)  / ggo.get(m, 1)) if ggo.get(m, 0) else None,
        'peso':        sf(mwg.get(m, 0)  / (mwg.get(m, 0) + ggo.get(m, 0)))
                       if (mwg.get(m, 0) + ggo.get(m, 0)) else None,
    })

# ── 11. montar payload final ──────────────────────────────────────────────────
print("→ Montando payload...")
RF = sum(r['Valor reembolsado'] for r in mensal)
CB = sum(r['Valor de chargeback'] for r in mensal)

P = {
    'mensal':       mensal,
    'coorte':       mensal,   # mesma fonte, lida de forma diferente no front
    'semanal':      [],       # deixar vazio — front usa cohort
    'cohort':       [{'Semana': r['sf'], 'Gross': r['g'], 'Mes': r['s'][:7],
                      **{f'W+{i}': r.get(f'W{i+1}') for i in range(13)}}
                     for r in coh_rf],
    'cohortV':      [],
    'decP':         dec_p,
    'decV':         dec_v,
    'produtos':     produtos,
    'bw':           [],
    'recon':        [],
    'cobprod':      [],
    'valid':        [],
    'custos':       [],
    'fases':        [],
    'afiliados':    af_rows,
    'motivos':      motivos,
    'cohortProduto': {p: [{'Semana': r['sf'], 'Mes': r['s'][:7], 'Gross': r['g'],
                            'Amostra': 'baixa (<$5k)' if r['g'] < 5000 else 'ok',
                            **{f'W+{i}': r.get(f'W{i+1}') for i in range(13)}}
                           for r in prf[p]] for p in prods},
    'cohortResumo': cohort_resumo,
    'cohortEx': {
        'rf':    coh_rf, 'cb': coh_cb,
        'prods': prods,  'prf': prf, 'pcb': pcb,
        'bw':    BW,
        'curvaA': curva_a, 'curvaB': curva_b,
    },
    'curvas': {
        'colsW': [f'W+{i}' for i in range(13)],
        'A': curva_a, 'B': curva_b,
    },
    'maxweb':    maxweb,
    'mwPerfil':  {},
    'mwAgosto':  {},
    'lipovive':  {},
    'tigerSemana': [],
    'sysSemana': sys_semana,
    'sysRef':    sys_ref,
    'pos17':     pos17,
    'salvamento': salvamento,
    'funil': {'recebidos': 187, 'resolvidos': 30, 'reembolsos': 12,
              'salvos': 11, 'diasSemana': 7, 'naoCBSystem': 82223.35,
              'retencaoObs': 0.5908},
    'gerado_em': str(DATA_REF.date()),
    # carimbo mostrado no rodapé do painel: quando o robô rodou e até que
    # dia o dado alcança. São coisas diferentes — o run das 3h traz o
    # movimento até o dia anterior.
    'atualizacao': {
        'rodou_em':   datetime.now(timezone.utc)
                        .astimezone(timezone(timedelta(hours=-3)))
                        .strftime('%d/%m/%Y às %H:%M'),
        'dado_ate':   DATA_REF.strftime('%d/%m/%Y'),
        'pedidos':    int(len(ped_a)),
        'reembolsos': int(len(ref_a)),
    },
}

# ── 12. injetar no template ───────────────────────────────────────────────────
print("→ Gerando HTML...")
if not TEMPLATE.exists():
    print(f"ERRO: template não encontrado em {TEMPLATE}")
    sys.exit(1)

template_html = TEMPLATE.read_text(encoding='utf-8')
payload_json  = json.dumps(P, ensure_ascii=False, separators=(',', ':'), default=str)
html          = template_html.replace('__PAYLOAD__', payload_json)

OUT.write_text(html, encoding='utf-8')
kb = OUT.stat().st_size // 1024
print(f"✓ {OUT}  ({kb} KB)  |  data ref: {DATA_REF.date()}")
